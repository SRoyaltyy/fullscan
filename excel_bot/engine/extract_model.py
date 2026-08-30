"""Extract a complete, machine-readable model from the Excel workbook.

Produces model.json containing:
  - formulas: every formula cell (incl. array formula text + spill ref)
  - cached:   every cached computed value (ground truth from Excel's last calc)
  - cf_rules: conditional formatting rules with resolved fill colors
  - static_fills: non-CF cell fills (headers etc.)
  - external: cached values of the external workbook link [1]
  - meta:     sheet dimensions, control cells
"""
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, date, time

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import column_index_from_string, get_column_letter

XLSX = "inputs/Simple View--Calculation.xlsx"
OUT = "engine/model.json"

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


# ---------------------------------------------------------------- theme ----
def load_theme_colors(zf):
    """Return list of 12 theme colors as RRGGBB (dk1,lt1,dk2,lt2,accent1-6,hlink,folHlink)."""
    xml = zf.read("xl/theme/theme1.xml")
    root = ET.fromstring(xml)
    a = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    scheme = root.find(f".//{a}clrScheme")
    colors = []
    for child in scheme:
        srgb = child.find(f"{a}srgbClr")
        sys = child.find(f"{a}sysClr")
        if srgb is not None:
            colors.append(srgb.get("val"))
        elif sys is not None:
            colors.append(sys.get("lastClr"))
        else:
            colors.append("000000")
    # Excel swaps dk1/lt1 and dk2/lt2 for theme indexing: 0=lt1? No:
    # theme color index: 0=dk1(text1)=windowText?, per OOXML the order in
    # clrMap decides. Standard Excel mapping: 0=lt1(bg1)=FFFFFF? Actually
    # Excel's theme color indices: 0=Background1(lt1),1=Text1(dk1),
    # 2=Background2(lt2),3=Text2(dk2),4..9=accent1..6
    # clrScheme order: dk1,lt1,dk2,lt2,accent1..6,hlink,folHlink
    if len(colors) >= 10:
        dk1, lt1, dk2, lt2 = colors[0], colors[1], colors[2], colors[3]
        rest = colors[4:10]
        return [lt1, dk1, lt2, dk2] + rest  # index 0..9
    return colors


def apply_tint(rgb, tint):
    """Apply Excel theme tint to an RRGGBB color."""
    if not tint:
        return rgb
    r, g, b = (int(rgb[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    def chan(c):
        if tint < 0:
            return c * (1.0 + tint)
        return c + (1.0 - c) * tint
    r, g, b = (chan(c) for c in (r, g, b))
    return "%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


# ----------------------------------------------------------------- dxfs ----
def load_dxf_fills(zf, theme):
    """dxfId -> resolved fill RGB ('RRGGBB') or None. Uses bgColor of solid fills."""
    xml = zf.read("xl/styles.xml")
    root = ET.fromstring(xml)
    dxfs = root.find(f"{{{NS}}}dxfs")
    fills = []
    if dxfs is None:
        return fills
    for dxf in dxfs:
        fill = dxf.find(f"{{{NS}}}fill")
        rgb = None
        if fill is not None:
            pf = fill.find(f"{{{NS}}}patternFill")
            if pf is not None:
                # Excel CF colors sit in bgColor for the solid CF patterns
                col = pf.find(f"{{{NS}}}bgColor")
                if col is None:
                    col = pf.find(f"{{{NS}}}fgColor")
                if col is not None:
                    if col.get("rgb") and col.get("rgb") != "00000000":
                        rgb = col.get("rgb")[-6:]
                    elif col.get("theme") is not None:
                        base = theme[int(col.get("theme"))]
                        rgb = apply_tint(base, float(col.get("tint") or 0))
                    elif col.get("indexed") is not None:
                        INDEXED = {2: "FF0000", 3: "00FF00", 5: "0000FF",
                                   10: "FF0000", 9: "FFFFFF", 8: "000000",
                                   7: "FFFF00", 64: "000000"}
                        rgb = INDEXED.get(int(col.get("indexed")), "000000")
        fills.append(rgb)
    return fills


# --------------------------------------------------------- external link ---
def load_external_cache(zf):
    """{(sheet_name, cell): value} from externalLink1.xml"""
    out = {}
    names = [n for n in zf.namelist() if re.match(r"xl/externalLinks/externalLink\d+\.xml$", n)]
    for n in names:
        root = ET.fromstring(zf.read(n))
        book = root.find(f"{{{NS}}}externalBook")
        sheet_names = [sn.get("val") for sn in book.find(f"{{{NS}}}sheetNames")]
        ds = book.find(f"{{{NS}}}sheetDataSet")
        if ds is None:
            continue
        for sd in ds:
            sid = int(sd.get("sheetId"))
            if sid >= len(sheet_names):
                continue
            sname = sheet_names[sid]
            for row in sd:
                for cell in row:
                    ref = cell.get("r")
                    t = cell.get("t")
                    v = cell.find(f"{{{NS}}}v")
                    if v is None or v.text is None:
                        val = ""
                    else:
                        txt = v.text
                        if t == "str":
                            val = txt
                        elif t == "b":
                            val = txt == "1"
                        elif t == "e":
                            val = txt  # error string
                        else:
                            try:
                                val = float(txt)
                                if val == int(val):
                                    val = int(val)
                            except ValueError:
                                val = txt
                    out[(sname, ref)] = val
    return out


# ------------------------------------------------------------------ main ---
def excel_serial(dt):
    """datetime/date -> Excel serial number (1900 system)."""
    epoch = datetime(1899, 12, 30)
    if isinstance(dt, datetime):
        delta = dt - epoch
    elif isinstance(dt, date):
        delta = datetime(dt.year, dt.month, dt.day) - epoch
    else:
        return dt
    v = delta.days + delta.seconds / 86400
    return int(v) if v == int(v) else v


def norm_cached(v):
    if isinstance(v, (datetime, date)):
        return excel_serial(v)
    if isinstance(v, time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400
    return v


def main():
    zf = zipfile.ZipFile(XLSX)
    theme = load_theme_colors(zf)
    dxf_fills = load_dxf_fills(zf, theme)
    external = load_external_cache(zf)

    wbf = openpyxl.load_workbook(XLSX)               # formulas
    wbv = openpyxl.load_workbook(XLSX, data_only=True)  # cached values
    wsf = wbf["Sheet1"]
    wsv = wbv["Sheet1"]

    formulas = {}   # coord -> {f, ref(optional)}
    cached = {}     # coord -> value
    static_fills = {}

    for row in wsf.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, ArrayFormula):
                formulas[c.coordinate] = {"f": v.text, "ref": str(v.ref)}
            elif isinstance(v, str) and v.startswith("="):
                formulas[c.coordinate] = {"f": v}
            # static fill (theme/rgb solid)
            fl = c.fill
            if fl is not None and fl.patternType == "solid":
                col = fl.fgColor
                rgb = None
                if col.type == "rgb" and isinstance(col.rgb, str):
                    rgb = col.rgb[-6:]
                elif col.type == "theme":
                    rgb = apply_tint(theme[col.theme], col.tint or 0)
                if rgb and rgb != "000000":
                    static_fills[c.coordinate] = rgb

    for row in wsv.iter_rows():
        for c in row:
            if c.value is not None:
                cached[c.coordinate] = norm_cached(c.value)

    # conditional formatting via raw XML (keeps dxfId + priority exact)
    cf_rules = []
    sheet_xml = zf.read("xl/worksheets/sheet1.xml")
    root = ET.fromstring(sheet_xml)
    for cf in root.iter(f"{{{NS}}}conditionalFormatting"):
        sqref = cf.get("sqref")
        for rule in cf.findall(f"{{{NS}}}cfRule"):
            dxf_id = rule.get("dxfId")
            entry = {
                "sqref": sqref,
                "type": rule.get("type"),
                "priority": int(rule.get("priority")),
                "operator": rule.get("operator"),
                "stopIfTrue": rule.get("stopIfTrue") == "1",
                "formulas": [f.text or "" for f in rule.findall(f"{{{NS}}}formula")],
                "fill": dxf_fills[int(dxf_id)] if dxf_id is not None and int(dxf_id) < len(dxf_fills) else None,
            }
            if rule.get("type") == "colorScale":
                cs = rule.find(f"{{{NS}}}colorScale")
                cfvos = [(c.get("type"), c.get("val")) for c in cs.findall(f"{{{NS}}}cfvo")]
                cols = []
                for c in cs.findall(f"{{{NS}}}color"):
                    if c.get("rgb"):
                        cols.append(c.get("rgb")[-6:])
                    elif c.get("theme") is not None:
                        cols.append(apply_tint(theme[int(c.get("theme"))], float(c.get("tint") or 0)))
                entry["colorScale"] = {"cfvo": cfvos, "colors": cols}
            cf_rules.append(entry)

    model = {
        "meta": {
            "sheet": "Sheet1",
            "max_row": wsf.max_row,
            "max_col": wsf.max_column,
            "theme": theme,
        },
        "formulas": formulas,
        "cached": cached,
        "cf_rules": sorted(cf_rules, key=lambda r: r["priority"]),
        "static_fills": static_fills,
        "external": {f"{s}!{c}": v for (s, c), v in external.items()},
    }
    with open(OUT, "w") as fh:
        json.dump(model, fh)
    print(f"formulas={len(formulas)} cached={len(cached)} cf_rules={len(cf_rules)} "
          f"external={len(external)} static_fills={len(static_fills)}")
    # show the two STOCKHISTORY anchors
    for coord in ("IR1", "AP1", "O1", "P1", "Q1", "R1"):
        print(coord, "formula:", formulas.get(coord), "| cached:", cached.get(coord))


if __name__ == "__main__":
    main()
