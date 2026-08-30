"""End-to-end runner: ticker + date -> per-day values & conditional colors.

Modes:
  python run.py --from-cache
      Replicate the workbook exactly as Excel last calculated it
      (uses the cached STOCKHISTORY data inside the file).
  python run.py --ticker AAPL [--date 2026-07-27]
      Fetch fresh data (Yahoo datasource) and run the full engine.

Outputs (in outputs/):
  grid_<tag>.html   visual replica of the Excel view (columns A-O)
  grid_<tag>.json   per-day values + fill colors (machine-readable)
"""
import argparse
import json
import os
import sys
from datetime import datetime, date, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl.utils import get_column_letter

from evaluator import Evaluator, split_coord
from colors import ColorEngine, serial_to_date
from xlrt import Err, is_arr
from validate import build_seeds
from stockhistory import stockhistory_array, serial

VISIBLE_COLS = list(range(1, 16))  # A..O
ROW_START, ROW_END = 1, 145


def seed_live(ev, ticker, run_date, workspace="."):
    """Seed STOCKHISTORY spill regions from freshly fetched data."""
    start_d = run_date - timedelta(days=200)
    start_w = run_date - timedelta(days=600)
    daily = stockhistory_array(ticker, start_d, run_date, 0, workspace)
    weekly = stockhistory_array(ticker, start_w, run_date, 1, workspace)
    seeds = {}
    for anchor_col, grid, maxr in (("IR", daily, 160), ("AP", weekly, 100)):
        c0 = split_coord(f"{anchor_col}1")[0]
        for i, rowvals in enumerate(grid):
            for j, v in enumerate(rowvals):
                seeds[f"{get_column_letter(c0 + j)}{1 + i}"] = v
        # blank out stale cells below the data
        for r in range(1 + len(grid), maxr + 1):
            for j in range(6):
                seeds[f"{get_column_letter(c0 + j)}{r}"] = None
    seeds["O1"] = ticker
    return seeds, len(daily) - 1, len(weekly) - 1


def fmt_value(v, col_idx):
    if isinstance(v, Err):
        return v.code
    if v is None:
        return ""
    if col_idx == 1 and isinstance(v, (int, float)) and v > 30000:
        try:
            d = serial_to_date(v)
            return f"{d.day}/{d.month}/{d.year}"
        except (OverflowError, ValueError):
            pass
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return f"{int(v):,}"
        return f"{v:,.4g}" if abs(v) < 100 else f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


def run(ticker=None, run_date=None, from_cache=False, tag=None, workspace="."):
    model = json.load(open(os.path.join(workspace, "engine/model.json")))
    today_serial = model["cached"].get("P1")
    if from_cache:
        run_date = serial_to_date(today_serial)
        ticker = model["cached"].get("O1")
    elif run_date is None:
        run_date = date.today()
    if isinstance(run_date, str):
        run_date = datetime.strptime(run_date, "%Y-%m-%d").date()
    today = serial(run_date)

    ev = Evaluator(os.path.join(workspace, "engine/model.json"), today=today)
    if from_cache:
        ev.seed(build_seeds(model))
        ndays = nweeks = None
    else:
        seeds, ndays, nweeks = seed_live(ev, ticker, run_date, workspace)
        ev.seed(seeds)
    ce = ColorEngine(ev, os.path.join(workspace, "engine/model.json"))

    grid = []
    for r in range(ROW_START, ROW_END + 1):
        row = []
        for c in VISIBLE_COLS:
            coord = f"{get_column_letter(c)}{r}"
            v = ev.get_cell(coord)
            if is_arr(v):
                v = v[0][0]
            fill = ce.fill_for(coord)
            row.append({"cell": coord, "value": None if isinstance(v, Err) else v,
                        "error": v.code if isinstance(v, Err) else None,
                        "text": fmt_value(v, c), "fill": fill})
        grid.append(row)

    tag = tag or (f"replica_{ticker}_{run_date}" if from_cache
                  else f"{ticker}_{run_date}")
    outdir = os.path.join(workspace, "outputs")
    os.makedirs(outdir, exist_ok=True)
    json_path = os.path.join(outdir, f"grid_{tag}.json")
    json.dump({"ticker": ticker, "date": str(run_date), "grid": grid},
              open(json_path, "w"), default=str)

    html_path = os.path.join(outdir, f"grid_{tag}.html")
    write_html(html_path, grid, ticker, run_date, from_cache, ndays, nweeks)
    return html_path, json_path


def write_html(path, grid, ticker, run_date, from_cache, ndays, nweeks):
    cols = "ABCDEFGHIJKLMNO"
    parts = [f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>{ticker} - {run_date}</title><style>
body {{ font-family: Calibri, Arial, sans-serif; background:#222; color:#eee; padding:16px; }}
table {{ border-collapse: collapse; font-size: 12px; }}
td {{ border: 1px solid #999; padding: 1px 6px; min-width: 64px; text-align: right;
     color: #111; background: #fff; white-space: nowrap; }}
td.h {{ background: #ddd; font-weight: bold; }}
.meta {{ margin-bottom: 10px; }}
</style></head><body>
<div class="meta"><b>{ticker}</b> as of {run_date}
({'Excel cached replica' if from_cache else f'live fetch: {ndays} daily rows, {nweeks} weekly rows'})</div>
<table>"""]
    for i, row in enumerate(grid):
        parts.append("<tr>")
        for cell in row:
            fill = cell["fill"]
            style = f' style="background:#{fill}"' if fill else ""
            cls = ' class="h"' if i == 0 else ""
            parts.append(f'<td{cls}{style}>{cell["text"]}</td>')
        parts.append("</tr>")
    parts.append("</table></body></html>")
    open(path, "w", encoding="utf-8").write("\n".join(parts))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker")
    ap.add_argument("--date")
    ap.add_argument("--from-cache", action="store_true")
    ap.add_argument("--tag")
    args = ap.parse_args()
    html, js = run(ticker=args.ticker, run_date=args.date,
                   from_cache=args.from_cache, tag=args.tag)
    print("HTML:", html)
    print("JSON:", js)
