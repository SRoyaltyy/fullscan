"""Backtest grid builder: per ticker, run the Excel engine over anchor dates
tiling the history, and save per-day OHLCV + 15-column color fills.

Design:
  - one Evaluator per anchor; anchor step = STEP trading days (<=136 coverage)
  - each anchor needs 200 calendar days daily + 600 days weekly lookback
  - resumable: skips tickers with existing grids/<TICKER>.json
  - chunked: --time-budget seconds, exits cleanly when budget is spent

Usage:
  python engine/backtest.py --tickers AAPL MSFT --anchors-from latest
  python engine/backtest.py --sample 40 --seed 7 --time-budget 260 --workers 1
"""
import argparse
import json
import os
import random
import sys
import time
from datetime import date, datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl.utils import get_column_letter

from evaluator import Evaluator
from colors import ColorEngine
from stockhistory import fetch_daily, weekly_aggregate, serial
from universe import load_universe

GRIDS_DIR = os.environ.get("GRIDS_DIR", "grids")
STEP = 120          # trading days between anchors
COVERAGE = 136      # trading days of colors per engine run
MODEL = "engine/model.json"

_MODEL_CACHE = {}


def get_model():
    if "m" not in _MODEL_CACHE:
        _MODEL_CACHE["m"] = json.load(open(MODEL))
    return _MODEL_CACHE["m"]


def stockhistory_from_rows(rows, start_d, end_d, interval):
    """STOCKHISTORY-equivalent array from prefetched daily rows."""
    sel = [r for r in rows if start_d <= r["date"] <= end_d]
    if interval == 1:
        sel = weekly_aggregate(sel)
    grid = [["Date", "Close", "Open", "High", "Low", "Volume"]]
    for r in sel:
        grid.append([serial(r["date"]), r["close"], r["open"],
                     r["high"], r["low"], r["volume"]])
    return grid


def seed_anchor(ev, ticker, rows, anchor):
    start_d, start_w = anchor - timedelta(days=200), anchor - timedelta(days=600)
    daily = stockhistory_from_rows(rows, start_d, anchor, 0)
    weekly = stockhistory_from_rows(rows, start_w, anchor, 1)
    seeds = {}
    from evaluator import split_coord
    for anchor_col, grid, maxr in (("IR", daily, 160), ("AP", weekly, 100)):
        c0 = split_coord(f"{anchor_col}1")[0]
        for i, rowvals in enumerate(grid):
            for j, v in enumerate(rowvals):
                seeds[f"{get_column_letter(c0 + j)}{1 + i}"] = v
        for r in range(1 + len(grid), maxr + 1):
            for j in range(6):
                seeds[f"{get_column_letter(c0 + j)}{r}"] = None
    seeds["O1"] = ticker
    return seeds


def run_anchor(ticker, rows, anchor):
    """One engine run -> list of day dicts (date, ohlcv, fills)."""
    ev = Evaluator(MODEL, today=serial(anchor))
    ev.seed(seed_anchor(ev, ticker, rows, anchor))
    ce = ColorEngine(ev, MODEL)
    days = []
    for r in range(2, 146):
        a = ev.get_cell(f"A{r}")
        if not isinstance(a, (int, float)) or a < 30000:
            continue
        vals = {}
        for name, col in (("close", "B"), ("open", "C"), ("high", "D"),
                          ("low", "E"), ("volume", "F")):
            v = ev.get_cell(f"{col}{r}")
            vals[name] = v if isinstance(v, (int, float)) else None
        fills = [ce.fill_for(f"{get_column_letter(c)}{r}") for c in range(1, 16)]
        days.append({"date": int(a), **vals, "fills": fills})
    return days


def build_ticker(ticker, rows, anchors):
    """Run all anchors, stitch days (newest STEP days per anchor), return table."""
    by_date = {}
    anchors = sorted(anchors)
    for k, anchor in enumerate(anchors):
        cutoff = anchors[k - 1] if k > 0 else None
        for d in run_anchor(ticker, rows, anchor):
            dt = (datetime(1899, 12, 30) + timedelta(days=d["date"])).date()
            if cutoff is not None and dt <= cutoff:
                continue
            by_date[dt] = d
    return [by_date[k] for k in sorted(by_date)]


def pick_anchors(rows, earliest, latest):
    """Anchor dates (trading days) so coverage tiles [earliest, latest]."""
    tdays = [r["date"] for r in rows if earliest <= r["date"] <= latest]
    if not tdays:
        return []
    anchors = [tdays[-1]]
    i = len(tdays) - 1 - STEP
    while i >= 0:
        anchors.append(tdays[i])
        i -= STEP
    return sorted(anchors)


def fetch_stitched(ticker, start, end):
    """Fetch potentially >2y of history via 729-day chunks (datasource limit),
    stitched and deduped. Chunks are cached on disk individually."""
    rows = {}
    chunk_start = start
    while chunk_start < end:
        chunk_end = min(chunk_start + timedelta(days=729), end)
        for r in fetch_daily(ticker, chunk_start, chunk_end):
            rows[r["date"]] = r
        chunk_start = chunk_end + timedelta(days=1)
    return [rows[k] for k in sorted(rows)]


def process_ticker(ticker, fetch_start, fetch_end, backtest_start):
    """Fetch + all anchors + save grid. Returns (ticker, n_days, error)."""
    try:
        rows = fetch_stitched(ticker, fetch_start, fetch_end)
        if len(rows) < 50:
            return (ticker, 0, "insufficient data")
        anchors = pick_anchors(rows, max(backtest_start, rows[0]["date"]),
                               rows[-1]["date"])
        if not anchors:
            return (ticker, 0, "no anchors")
        days = build_ticker(ticker, rows, anchors)
        if not days:
            return (ticker, 0, "no days")
        os.makedirs(GRIDS_DIR, exist_ok=True)
        json.dump({"ticker": ticker, "days": days},
                  open(os.path.join(GRIDS_DIR, f"{ticker}.json"), "w"))
        return (ticker, len(days), None)
    except Exception as e:  # noqa: BLE001
        return (ticker, 0, f"{type(e).__name__}: {e}")


def _work(t):
    ticker, fetch_start, fetch_end, backtest_start = t
    return process_ticker(ticker, fetch_start, fetch_end, backtest_start)


FAIL_LOG = os.path.join(GRIDS_DIR, "_failed.json")


def load_failed():
    if os.path.exists(FAIL_LOG):
        return json.load(open(FAIL_LOG))
    return {}


LOCK = os.path.join(GRIDS_DIR, "_lock")


def acquire_lock():
    """Single-chunk lock; returns True if acquired. Stale locks (>15 min)
    are broken so a crashed chunk never wedges the build."""
    try:
        fd = os.open(LOCK, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(time.time()).encode())
        os.close(fd)
        return True
    except FileExistsError:
        try:
            if time.time() - float(open(LOCK).read() or 0) > 900:
                os.remove(LOCK)
                return acquire_lock()
        except (ValueError, OSError):
            pass
        return False


def release_lock():
    try:
        os.remove(LOCK)
    except OSError:
        pass


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tickers", nargs="*")
    ap.add_argument("--sample", type=int)
    ap.add_argument("--seed", type=int, default=7)
    ap.add_argument("--time-budget", type=int, default=260)
    ap.add_argument("--workers", type=int, default=1)
    ap.add_argument("--deep-years", type=int, default=0,
                    help="fetch & backtest N years of history (stitched calls)")
    ap.add_argument("--span-days", type=int, default=130,
                    help="backtest span within the 2y fetch (rest is lookback)")
    args = ap.parse_args()

    t0 = time.time()
    if args.tickers is None and not acquire_lock():
        print("LOCKED")
        return
    try:
        _main_inner(args, t0)
    finally:
        if args.tickers is None:
            release_lock()


def _main_inner(args, t0):
    today = date.today()
    fetch_end = today
    if getattr(args, "deep_years", 0):
        fetch_start = today - timedelta(days=args.deep_years * 365)
        backtest_start = fetch_start + timedelta(days=600)
    else:
        fetch_start = today - timedelta(days=729)
        backtest_start = today - timedelta(days=args.span_days)

    if args.tickers:
        tickers = args.tickers
    else:
        uni = [u["symbol"] for u in load_universe()]
        random.seed(args.seed)
        random.shuffle(uni)
        done = {f[:-5] for f in os.listdir(GRIDS_DIR)} if os.path.isdir(GRIDS_DIR) else set()
        failed = load_failed()
        skip = done | {t for t, n in failed.items() if n >= 2}
        todo = [t for t in uni if t not in skip]
        print(f"universe todo={len(todo)} done={len(done)} failed>={sum(1 for t,n in failed.items() if n>=2)}")
        if not todo:
            print("COMPLETE")
            return
        tickers = todo[:args.sample or 40]

    jobs = [(t, fetch_start, fetch_end, backtest_start) for t in tickers]
    results = []
    if args.workers > 1:
        from multiprocessing import Pool
        with Pool(args.workers) as pool:
            for r in pool.imap_unordered(_work, jobs):
                results.append(r)
                print(f"{'OK ' if r[2] is None else 'ERR'} {r[0]:8s} days={r[1]:4d} "
                      f"{(r[2] or '')[:80]}", flush=True)
                if time.time() - t0 > args.time_budget:
                    pool.terminate()
                    print(f"[budget] stopping after {len(results)} tickers")
                    break
    else:
        for j in jobs:
            if time.time() - t0 > args.time_budget:
                print(f"[budget] stopping after {len(results)} tickers")
                break
            r = _work(j)
            results.append(r)
            print(f"{'OK ' if r[2] is None else 'ERR'} {r[0]:8s} days={r[1]:4d} "
                  f"{(r[2] or '')[:80]}", flush=True)

    failed = load_failed()
    for t, _, err in results:
        if err:
            failed[t] = failed.get(t, 0) + 1
    os.makedirs(GRIDS_DIR, exist_ok=True)
    json.dump(failed, open(FAIL_LOG, "w"))
    ok = sum(1 for r in results if r[2] is None)
    print(f"done: {ok}/{len(results)} ok, elapsed {time.time()-t0:.0f}s")


if __name__ == "__main__":
    main()
