"""Conditional-formatting color engine.

Reproduces Excel's CF evaluation:
  - rules sorted by priority; first matching rule supplies the fill
  - cellIs / expression / containsText / colorScale supported
  - expression formulas are re-anchored: relative refs shift by the cell's
    offset from the sqref top-left (Excel CF semantics)
"""
import json
import re

from openpyxl.utils import column_index_from_string, get_column_letter

from evaluator import Evaluator, Ctx
from xlrt import Err, is_err, compare, to_num
from xlparse import parse


def shift_refs(node, drow, dcol):
    """Return a copy of AST with relative refs shifted (CF re-anchoring)."""
    t = node[0]
    if t == "ref":
        _, sheet, ext, col, row, abscol, absrow = node
        ci = column_index_from_string(col)
        if not abscol:
            ci += dcol
        if not absrow:
            row += drow
        if ci < 1 or row < 1:
            return ("err", "#REF!")
        return ("ref", sheet, ext, get_column_letter(ci), row, abscol, absrow)
    if t == "range":
        return ("range", shift_refs(node[1], drow, dcol), shift_refs(node[2], drow, dcol))
    if t in ("bin",):
        return (t, node[1], shift_refs(node[2], drow, dcol), shift_refs(node[3], drow, dcol))
    if t in ("un", "pct"):
        return (t, node[1], shift_refs(node[2], drow, dcol))
    if t == "call":
        return (t, node[1], [shift_refs(a, drow, dcol) for a in node[2]])
    if t == "array":
        return (t, [[shift_refs(x, drow, dcol) for x in row] for row in node[1]])
    return node


def parse_sqref(sqref):
    """'I1:I145 I156:I200' or 'G2' -> list of (c1, r1, c2, r2)."""
    out = []
    for part in sqref.split():
        part = part.replace("$", "")
        if ":" in part:
            a, b = part.split(":")
        else:
            a = b = part
        m1 = re.match(r"([A-Z]+)(\d+)", a)
        m2 = re.match(r"([A-Z]+)(\d+)", b)
        out.append((column_index_from_string(m1.group(1)), int(m1.group(2)),
                    column_index_from_string(m2.group(1)), int(m1 and m2.group(2))))
    return out


class ColorEngine:
    def __init__(self, ev: Evaluator, model_path="engine/model.json"):
        self.ev = ev
        model = json.load(open(model_path))
        self.static_fills = model["static_fills"]
        self.ast_cache = {}
        self.const_cache = {}
        # precompute: per-column rule lists (global priority order preserved),
        # with sqrefs parsed once
        self.col_rules = {}
        for rule in model["cf_rules"]:
            ranges = parse_sqref(rule["sqref"])
            rule["_ranges"] = ranges
            cols = set()
            for (c1, r1, c2, r2) in ranges:
                cols.update(range(c1, min(c2, c1 + 300) + 1))
            for c in cols:
                self.col_rules.setdefault(c, []).append(rule)

    def fill_for(self, coord):
        """Resolved fill RRGGBB for a cell, or None (no fill)."""
        m = re.match(r"([A-Z]+)(\d+)", coord)
        col_idx, row = column_index_from_string(m.group(1)), int(m.group(2))
        for rule in self.col_rules.get(col_idx, ()):
            hit = None
            for (c1, r1, c2, r2) in rule["_ranges"]:
                if c1 <= col_idx <= c2 and r1 <= row <= r2:
                    hit = (c1, r1)
                    break
            if hit is None:
                continue
            if self._matches(rule, coord, col_idx, row, hit[0], hit[1]):
                if rule.get("fill"):
                    return rule["fill"]
                if rule.get("type") == "colorScale":
                    return self._color_scale(rule, coord)
                return None  # matched rule defines no fill
        return self.static_fills.get(coord)

    # ------------------------------------------------------------- rules ---
    def _matches(self, rule, coord, col_idx, row, anchor_c, anchor_r):
        rtype = rule["type"]
        val = self.ev.get_cell(coord)
        if rtype == "cellIs":
            op = rule["operator"]
            f1 = self._eval_const(rule["formulas"][0])
            if isinstance(f1, Err):
                return False
            if op == "between":
                f2 = self._eval_const(rule["formulas"][1])
                lo, hi = (f1, f2) if f1 <= f2 else (f2, f1)
                return compare(">=", val, lo) is True and compare("<=", val, hi) is True
            if op == "notBetween":
                f2 = self._eval_const(rule["formulas"][1])
                lo, hi = (f1, f2) if f1 <= f2 else (f2, f1)
                return not (compare(">=", val, lo) is True and compare("<=", val, hi) is True)
            opmap = {"greaterThan": ">", "greaterThanOrEqual": ">=",
                     "lessThan": "<", "lessThanOrEqual": "<=",
                     "equal": "=", "notEqual": "<>"}
            return compare(opmap[op], val, f1) is True
        if rtype in ("expression", "containsText"):
            res = self._eval_expr(rule["formulas"][0], col_idx - anchor_c,
                                  row - anchor_r, coord)
            return res is True
        if rtype == "colorScale":
            return isinstance(val, (int, float)) and not isinstance(val, bool)
        return False

    def _eval_const(self, text):
        if text not in self.const_cache:
            self.const_cache[text] = self.ev.eval_node(parse(text), Ctx(self.ev))
        return self.const_cache[text]

    def _eval_expr(self, text, dcol, drow, coord):
        ast = self.ast_cache.get(text)
        if ast is None:
            ast = parse(text)
            self.ast_cache[text] = ast
        m = re.match(r"([A-Z]+)(\d+)", coord)
        cell = (int(m.group(2)), column_index_from_string(m.group(1)))
        return self.ev.eval_node(ast, Ctx(self.ev, cell=cell, off=(drow, dcol)))

    def _color_scale(self, rule, coord):
        cs = rule["colorScale"]
        val = self.ev.get_cell(coord)
        if not isinstance(val, (int, float)) or isinstance(val, bool):
            return None
        cfvo, colors = cs["cfvo"], cs["colors"]
        # gather the value domain from the rule's first sqref range
        c1, r1, c2, r2 = parse_sqref(rule["sqref"])[0]
        vals = []
        for rr in range(r1, min(r2, 10000) + 1):
            for cc in range(c1, c2 + 1):
                v = self.ev.get_cell(f"{get_column_letter(cc)}{rr}")
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    vals.append(v)
        if not vals:
            return None
        points = []
        for typ, raw in cfvo:
            if typ == "min":
                points.append(min(vals))
            elif typ == "max":
                points.append(max(vals))
            elif typ == "percentile":
                s = sorted(vals)
                k = float(raw) / 100 * (len(s) - 1)
                lo = int(k)
                points.append(s[lo] + (s[min(lo + 1, len(s) - 1)] - s[lo]) * (k - lo))
            elif typ == "percent":
                points.append(min(vals) + (max(vals) - min(vals)) * float(raw) / 100)
            else:
                points.append(float(raw))
        # interpolate
        if val <= points[0]:
            return colors[0]
        if val >= points[-1]:
            return colors[-1]
        for i in range(len(points) - 1):
            if points[i] <= val <= points[i + 1]:
                span = points[i + 1] - points[i]
                t = 0 if span == 0 else (val - points[i]) / span
                c1x = tuple(int(colors[i][k:k + 2], 16) for k in (0, 2, 4))
                c2x = tuple(int(colors[i + 1][k:k + 2], 16) for k in (0, 2, 4))
                mix = tuple(round(a + (b - a) * t) for a, b in zip(c1x, c2x))
                return "%02X%02X%02X" % mix
        return colors[-1]


def serial_to_date(n):
    from datetime import datetime, timedelta
    return (datetime(1899, 12, 30) + timedelta(days=float(n))).date()


if __name__ == "__main__":
    import sys
    sys.path.insert(0, "engine")
    from validate import build_seeds
    model = json.load(open("engine/model.json"))
    ev = Evaluator("engine/model.json", today=model["cached"].get("P1"))
    ev.seed(build_seeds(model))
    ce = ColorEngine(ev)
    # sample: columns A-O, rows 105-135 like the screenshot
    for r in [105, 110, 115, 120, 128, 130, 135]:
        line = []
        for c in range(1, 16):
            coord = f"{get_column_letter(c)}{r}"
            f = ce.fill_for(coord)
            line.append(f"{coord}:{f or '------'}")
        print(" ".join(line))
