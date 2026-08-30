"""Lazy grid evaluator + Excel function library.

Evaluates the extracted model (model.json) with:
  - memoized per-cell evaluation and cycle detection
  - array/spill semantics (anchor ref coverage)
  - external workbook cache resolution  [1]Change!X1
  - STOCKHISTORY interception via a pluggable data provider
  - injectable TODAY() for reproducible runs
"""
import json
import math
import re
import sys

sys.setrecursionlimit(300000)

from xlparse import parse, ParseError
from xlrt import (Err, E_VALUE, E_REF, E_NA, E_DIV0, E_NAME, E_NUM, E_CALC,
                  ERROR_STRINGS, is_err, is_arr, arr_shape, flatten, to_array,
                  first, to_num, to_text, to_bool, compare, broadcast_bin,
                  elementwise_un, make_criteria, _wildcard_re, _eq, _sort_key)

from openpyxl.utils import column_index_from_string, get_column_letter


class CycleError(Exception):
    pass


class XLLambda:
    __slots__ = ("params", "body", "env")

    def __init__(self, params, body, env):
        self.params = params
        self.body = body
        self.env = dict(env)


class Ctx:
    __slots__ = ("ev", "env", "cell", "off")

    def __init__(self, ev, env=None, cell=None, off=(0, 0)):
        self.ev = ev
        self.env = env if env is not None else {}
        self.cell = cell  # (row, col_index) of the cell being evaluated
        self.off = off    # (drow, dcol) applied to relative refs (CF anchoring)


class Evaluator:
    def __init__(self, model_path="engine/model.json", today=None,
                 stock_provider=None):
        model = json.load(open(model_path))
        self.formulas = model["formulas"]
        self.cached = model["cached"]
        self.external = model["external"]
        self.max_row = model["meta"]["max_row"]
        self.today = today
        self.stock_provider = stock_provider
        self.memo = {}
        self.in_progress = set()
        self.ast_cache = {}
        self.parse_errors = {}
        # array anchor coverage: coord -> (anchor_coord, drow, dcol)
        self.covered = {}
        for coord, f in self.formulas.items():
            ref = f.get("ref")
            if ref and ":" in ref:
                a, b = ref.split(":")
                c1, r1 = split_coord(a)
                c2, r2 = split_coord(b)
                if (c1, r1) != (c2, r2):
                    for rr in range(r1, r2 + 1):
                        for cc in range(c1, c2 + 1):
                            co = f"{get_column_letter(cc)}{rr}"
                            if co != coord:
                                self.covered[co] = (coord, rr - r1, cc - c1)

    # ------------------------------------------------------------ cells ----
    def seed(self, coord_values):
        """Pre-populate memo (e.g. cached STOCKHISTORY spill for validation)."""
        self.memo.update(coord_values)

    def get_cell(self, coord, depth=0):
        if coord in self.memo:
            return self.memo[coord]
        if coord in self.in_progress:
            raise CycleError(coord)
        # covered by an array anchor?
        cov = self.covered.get(coord)
        if cov and coord not in self.formulas:
            anchor, dr, dc = cov
            arr = self.get_cell(anchor)
            if is_arr(arr):
                try:
                    v = arr[dr][dc]
                except IndexError:
                    v = E_NA
            else:
                v = arr if (dr == 0 and dc == 0) else E_NA
            self.memo[coord] = v
            return v
        f = self.formulas.get(coord)
        if f is None:
            v = self.cached.get(coord)  # literal cell
            v = norm_literal(v)
            self.memo[coord] = v
            return v
        self.in_progress.add(coord)
        try:
            v = self.eval_formula(coord, f["f"])
        except CycleError:
            raise
        except Exception as e:  # noqa: BLE001 - record and surface as error
            v = Err(f"#PYERR:{type(e).__name__}:{e}")
        finally:
            self.in_progress.discard(coord)
        # single-cell formula displaying an array -> top-left
        if is_arr(v):
            ref = f.get("ref")
            if not ref or ":" not in ref:
                v = first(v)
        if v is None:
            v = 0.0  # Excel: formula result of empty reference displays as 0
        self.memo[coord] = v
        return v

    def eval_formula(self, coord, text):
        ast = self.ast_cache.get(coord)
        if ast is None:
            try:
                ast = parse(text)
            except ParseError as e:
                if text.strip("= \n") == "":
                    # junk cell containing only "="; Excel shows it as text
                    return self.cached.get(coord, "=")
                self.parse_errors[coord] = str(e)
                return Err(f"#PARSE:{e}")
            self.ast_cache[coord] = ast
        row = int(re.match(r"[A-Z]+(\d+)", coord).group(1))
        col = column_index_from_string(re.match(r"([A-Z]+)", coord).group(1))
        return self.eval_node(ast, Ctx(self, cell=(row, col)))

    # ------------------------------------------------------------ ranges ---
    def get_range(self, sheet, ext, c1, r1, c2, r2):
        if ext is not None:
            return self.get_external_range(sheet, c1, r1, c2, r2)
        if sheet not in (None, "Sheet1"):
            return E_REF
        out = []
        for rr in range(r1, r2 + 1):
            row = []
            for cc in range(c1, c2 + 1):
                row.append(self.get_cell(f"{get_column_letter(cc)}{rr}"))
            out.append(row)
        return out

    def get_external_cell(self, sheet, col_letter, row):
        v = self.external.get(f"{sheet}!{col_letter}{row}")
        return norm_literal(v)

    def get_external_range(self, sheet, c1, r1, c2, r2):
        return [[self.get_external_cell(sheet, get_column_letter(cc), rr)
                 for cc in range(c1, c2 + 1)] for rr in range(r1, r2 + 1)]

    # ----------------------------------------------------- ref resolution --
    def resolve_coords(self, node, ctx):
        """Resolve a node to (ext, sheet, col_idx, row) when it denotes a cell
        reference — handles plain refs and INDEX(range, n) reference form."""
        if node[0] == "ref":
            _, sheet, ext, col, row, abscol, absrow = node
            ci = column_index_from_string(col)
            if ctx.off != (0, 0):
                if not abscol:
                    ci += ctx.off[1]
                if not absrow:
                    row += ctx.off[0]
            return (ext, sheet, ci, row)
        if node[0] == "call" and node[1] == "INDEX":
            base = self.range_bbox(node[2][0], ctx)
            if base is None:
                return None
            ext, sheet, c1, r1, c2, r2 = base
            rownum = self.eval_node(node[2][1], ctx) if len(node[2]) > 1 else 1
            if isinstance(rownum, Err):
                return None
            rownum = int(to_num(rownum))
            colnum = 1
            if len(node[2]) > 2:
                colnum = self.eval_node(node[2][2], ctx)
                if isinstance(colnum, Err):
                    return None
                colnum = int(to_num(colnum))
            if c1 == c2:  # column vector: index down
                return (ext, sheet, c1, r1 + rownum - 1)
            if r1 == r2 and len(node[2]) == 2:  # row vector: index across
                return (ext, sheet, c1 + rownum - 1, r1)
            return (ext, sheet, c1 + colnum - 1, r1 + rownum - 1)
        return None

    def range_bbox(self, node, ctx):
        """Bounding box (ext, sheet, c1, r1, c2, r2) for range-like nodes."""
        if node[0] == "ref":
            _, sheet, ext, col, row, _, _ = node
            ci = column_index_from_string(col)
            return (ext, sheet, ci, row, ci, row)
        if node[0] == "range":
            a = self.resolve_coords(node[1], ctx)
            b = self.resolve_coords(node[2], ctx)
            if a is None or b is None:
                return None
            return (a[0], a[1], min(a[2], b[2]), min(a[3], b[3]),
                    max(a[2], b[2]), max(a[3], b[3]))
        if node[0] == "colrange":
            c1 = column_index_from_string(node[1])
            c2 = column_index_from_string(node[2])
            return (None, None, min(c1, c2), 1, max(c1, c2), 1048576)
        if node[0] == "rowrange":
            return (None, None, 1, min(node[1], node[2]), 16384,
                    max(node[1], node[2]))
        if node[0] == "extcolrange":
            _, ext, sheet, c1l, c2l = node
            c1, c2 = column_index_from_string(c1l), column_index_from_string(c2l)
            return (ext, sheet, min(c1, c2), 1, max(c1, c2), 1048576)
        return None

    # -------------------------------------------------------------- eval ---
    def eval_node(self, node, ctx):
        t = node[0]
        if t == "num":
            return node[1]
        if t == "str":
            return node[1]
        if t == "bool":
            return node[1]
        if t == "err":
            return ERROR_STRINGS.get(node[1], Err(node[1]))
        if t == "empty":
            return None
        if t == "name":
            v = ctx.env.get(node[1], None)
            if v is None and node[1] not in ctx.env:
                return E_NAME
            return v
        if t == "ref":
            _, sheet, ext, col, row, abscol, absrow = node
            ci = column_index_from_string(col)
            if ctx.off != (0, 0):
                if not abscol:
                    ci += ctx.off[1]
                if not absrow:
                    row += ctx.off[0]
                if ci < 1 or row < 1:
                    return E_REF
                col = get_column_letter(ci)
            if ext is not None:
                return self.get_external_cell(sheet, col, row)
            if sheet not in (None, "Sheet1"):
                return E_REF
            return self.get_cell(f"{col}{row}")
        if t == "range":
            a, b = node[1], node[2]
            ca = self.resolve_coords(a, ctx)
            cb = self.resolve_coords(b, ctx)
            if ca is None or cb is None:
                return E_REF
            ext, sheet, c1, r1 = ca
            _, _, c2, r2 = cb
            return self.get_range(sheet, ext, min(c1, c2), min(r1, r2),
                                  max(c1, c2), max(r1, r2))
        if t == "colrange":
            c1, c2 = column_index_from_string(node[1]), column_index_from_string(node[2])
            return self.get_range(None, None, min(c1, c2), 1, max(c1, c2), self.max_row)
        if t == "rowrange":
            r1, r2 = min(node[1], node[2]), max(node[1], node[2])
            return self.get_range(None, None, 1, r1, self.max_col(), r2)
        if t == "extcolrange":
            _, ext, sheet, c1l, c2l = node
            hi = self.external_max_row(sheet)
            return self.get_external_range(sheet, column_index_from_string(c1l), 1,
                                           column_index_from_string(c2l), hi)
        if t == "array":
            return [[self.eval_node(x, ctx) for x in row] for row in node[1]]
        if t == "un":
            v = self.eval_node(node[2], ctx)
            def neg(x):
                n = to_num(x)
                return n if isinstance(n, Err) else (-n if node[1] == "-" else n)
            return elementwise_un(neg, v)
        if t == "pct":
            v = self.eval_node(node[1], ctx)
            return elementwise_un(lambda x: (lambda n: n if isinstance(n, Err) else n / 100)(to_num(x)), v)
        if t == "bin":
            return self.eval_bin(node, ctx)
        if t == "call":
            return self.eval_call(node[1], node[2], ctx)
        return Err(f"#UNKNOWN_NODE:{t}")

    def max_col(self):
        return 275

    def external_max_row(self, sheet):
        hi = 1
        prefix = f"{sheet}!"
        for key in self.external:
            if key.startswith(prefix):
                m = re.search(r"(\d+)$", key)
                if m:
                    hi = max(hi, int(m.group(1)))
        return hi

    def eval_bin(self, node, ctx):
        _, op, ln, rn = node
        l = self.eval_node(ln, ctx)
        r = self.eval_node(rn, ctx)
        if op in ("=", "<>", "<", "<=", ">", ">="):
            return broadcast_bin(lambda a, b: compare(op, a, b), l, r)
        if op == "&":
            def concat2(a, b):
                x, y = to_text(a), to_text(b)
                if isinstance(x, Err):
                    return x
                if isinstance(y, Err):
                    return y
                return x + y
            return broadcast_bin(concat2, l, r)
        # arithmetic
        def arith(a, b):
            x, y = to_num(a), to_num(b)
            if isinstance(x, Err):
                return x
            if isinstance(y, Err):
                return y
            if op == "+":
                return x + y
            if op == "-":
                return x - y
            if op == "*":
                return x * y
            if op == "/":
                return E_DIV0 if y == 0 else x / y
            if op == "^":
                try:
                    if x == 0 and y < 0:
                        return E_DIV0
                    v = x ** y
                    if isinstance(v, complex):
                        return E_NUM
                    return v
                except (OverflowError, ZeroDivisionError):
                    return E_NUM
            return E_VALUE
        return broadcast_bin(arith, l, r)

    # -------------------------------------------------------------- calls --
    def eval_call(self, name, args, ctx):
        fn = SPECIAL.get(name)
        if fn:
            return fn(self, args, ctx)
        impl = FUNCS.get(name)
        if impl is None:
            return Err(f"#NAME?({name})")
        vals = [self.eval_node(a, ctx) for a in args]
        try:
            return impl(self, vals, ctx)
        except Exception as e:  # noqa: BLE001
            return Err(f"#FNERR:{name}:{e}")


def norm_literal(v):
    if isinstance(v, str) and v in ERROR_STRINGS:
        return ERROR_STRINGS[v]
    return v


def split_coord(coord):
    m = re.match(r"([A-Z]+)(\d+)", coord)
    return column_index_from_string(m.group(1)), int(m.group(2))


# ===========================================================================
#  function library
# ===========================================================================

def lift(fn):
    """Make a scalar function elementwise over array args (broadcast)."""
    def wrapper(ev, vals, ctx):
        if any(is_arr(v) for v in vals):
            shape = None
            for v in vals:
                if is_arr(v):
                    s = arr_shape(v)
                    shape = s if shape is None or s[0] * s[1] > shape[0] * shape[1] else shape
            rows, cols = shape
            out = []
            for i in range(rows):
                row = []
                for j in range(cols):
                    args = []
                    for v in vals:
                        if is_arr(v):
                            r, c = arr_shape(v)
                            args.append(v[i][j] if i < r and j < c else E_NA)
                        else:
                            args.append(v)
                    row.append(fn(*args))
                out.append(row)
            return out
        return fn(*vals)
    return wrapper


def nums_only(vals):
    """Collect numeric values from function args (ranges skip text/bool/blank)."""
    out = []
    for v in vals:
        if is_arr(v):
            for x in flatten(v):
                if isinstance(x, Err):
                    return x
                if isinstance(x, (int, float)) and not isinstance(x, bool):
                    out.append(float(x))
        else:
            if isinstance(v, Err):
                return v
            if v is None:
                continue
            n = to_num(v)
            if isinstance(n, Err):
                return n
            out.append(n)
    return out


def _flat1d(v):
    """Flatten a 1-column or 1-row array to a list."""
    if not is_arr(v):
        return [v]
    r, c = arr_shape(v)
    if c == 1:
        return [row[0] for row in v]
    if r == 1:
        return list(v[0])
    return None  # genuinely 2D


# ------------------------------------------------------------------ impls --

def f_sum(ev, vals, ctx):
    ns = nums_only(vals)
    return ns if isinstance(ns, Err) else sum(ns)


def f_average(ev, vals, ctx):
    ns = nums_only(vals)
    if isinstance(ns, Err):
        return ns
    return E_DIV0 if not ns else sum(ns) / len(ns)


def f_min(ev, vals, ctx):
    ns = nums_only(vals)
    if isinstance(ns, Err):
        return ns
    return min(ns) if ns else 0.0


def f_max(ev, vals, ctx):
    ns = nums_only(vals)
    if isinstance(ns, Err):
        return ns
    return max(ns) if ns else 0.0


def f_median(ev, vals, ctx):
    ns = nums_only(vals)
    if isinstance(ns, Err):
        return ns
    if not ns:
        return E_DIV0
    ns = sorted(ns)
    n = len(ns)
    mid = n // 2
    return ns[mid] if n % 2 else (ns[mid - 1] + ns[mid]) / 2


def _stdev(vals, sample):
    ns = nums_only(vals)
    if isinstance(ns, Err):
        return ns
    n = len(ns)
    if n < (2 if sample else 1):
        return E_DIV0
    mean = sum(ns) / n
    var = sum((x - mean) ** 2 for x in ns) / (n - 1 if sample else n)
    return math.sqrt(var)


def f_stdevp(ev, vals, ctx):
    return _stdev(vals, False)


def f_stdev(ev, vals, ctx):
    return _stdev(vals, True)


def f_sqrt(ev, vals, ctx):
    return lift(lambda x: (lambda n: n if isinstance(n, Err) else (math.sqrt(n) if n >= 0 else E_NUM))(to_num(x)))(ev, vals, ctx)


def f_abs(ev, vals, ctx):
    return lift(lambda x: (lambda n: n if isinstance(n, Err) else abs(n))(to_num(x)))(ev, vals, ctx)


def _agg_logical(vals, want_all):
    saw = False
    first_err = None
    result = want_all
    for v in vals:
        items = flatten(v) if is_arr(v) else [v]
        for x in items:
            if isinstance(x, Err):
                if first_err is None:
                    first_err = x
                continue
            if x is None or isinstance(x, str):
                if is_arr(v):
                    continue  # ranges ignore text/blank
                if isinstance(x, str):
                    b = to_bool(x)
                    if isinstance(b, Err):
                        if first_err is None:
                            first_err = b
                        continue
                    saw = True
                    result = (result and b) if want_all else (result or b)
                continue
            b = bool(x)
            saw = True
            result = (result and b) if want_all else (result or b)
    if first_err is not None:
        return first_err  # Excel AND/OR propagate errors (no short-circuit)
    if not saw:
        return E_VALUE
    return result


def f_and(ev, vals, ctx):
    return _agg_logical(vals, True)


def f_or(ev, vals, ctx):
    return _agg_logical(vals, False)


def f_isnumber(ev, vals, ctx):
    return lift(lambda x: isinstance(x, (int, float)) and not isinstance(x, bool))(ev, vals, ctx)


def f_iserror(ev, vals, ctx):
    return lift(lambda x: isinstance(x, Err))(ev, vals, ctx)


def f_isodd(ev, vals, ctx):
    return lift(lambda x: (lambda n: n if isinstance(n, Err) else (int(abs(n)) % 2 == 1))(to_num(x)))(ev, vals, ctx)


def f_isblank(ev, vals, ctx):
    return lift(lambda x: x is None)(ev, vals, ctx)


def f_len(ev, vals, ctx):
    return lift(lambda x: (lambda t: t if isinstance(t, Err) else len(t))(to_text(x)))(ev, vals, ctx)


def f_exact(ev, vals, ctx):
    return lift(lambda a, b: (lambda x, y: x == y if not isinstance(x, Err) and not isinstance(y, Err) else (x if isinstance(x, Err) else y))(to_text(a), to_text(b)))(ev, vals, ctx)


def f_mid(ev, vals, ctx):
    def mid(t, s, n):
        t = to_text(t)
        s, n = to_num(s), to_num(n)
        if isinstance(t, Err):
            return t
        if isinstance(s, Err):
            return s
        if isinstance(n, Err):
            return n
        s = int(s)
        return t[max(s - 1, 0):max(s - 1, 0) + int(n)]
    return lift(mid)(ev, vals, ctx)


def f_find(ev, vals, ctx):
    def find(sub, text, *rest):
        sub, text = to_text(sub), to_text(text)
        if isinstance(sub, Err):
            return sub
        if isinstance(text, Err):
            return text
        start = int(to_num(rest[0])) if rest else 1
        i = text.find(sub, start - 1)
        return E_VALUE if i < 0 else i + 1
    return lift(find)(ev, vals, ctx)


def f_search(ev, vals, ctx):
    def search(sub, text, *rest):
        sub, text = to_text(sub), to_text(text)
        if isinstance(sub, Err):
            return sub
        if isinstance(text, Err):
            return text
        start = int(to_num(rest[0])) if rest else 1
        body = _wildcard_re(sub).pattern[1:-1]  # strip ^ and $
        pat = re.compile(body, re.IGNORECASE)
        m = pat.search(text, start - 1)
        return E_VALUE if not m else m.start() + 1
    return lift(search)(ev, vals, ctx)


def f_concat(ev, vals, ctx):
    out = []
    for v in vals:
        for x in flatten(v):
            t = to_text(x)
            if isinstance(t, Err):
                return t
            out.append(t)
    return "".join(out)


def f_textjoin(ev, vals, ctx):
    if len(vals) < 3:
        return E_VALUE
    delim = to_text(vals[0])
    ignore_empty = to_bool(vals[1])
    parts = []
    for v in vals[2:]:
        for x in flatten(v):
            t = to_text(x)
            if isinstance(t, Err):
                return t
            if ignore_empty and t == "":
                continue
            parts.append(t)
    return delim.join(parts)


def f_textsplit(ev, vals, ctx):
    text = to_text(vals[0])
    if isinstance(text, Err):
        return text
    col_delim = to_text(vals[1]) if len(vals) > 1 and vals[1] is not None else None
    row_delim = to_text(vals[2]) if len(vals) > 2 and vals[2] is not None else None
    rows = [text] if row_delim is None else text.split(row_delim)
    grid = []
    for r in rows:
        grid.append([r] if col_delim is None else r.split(col_delim))
    return grid


def f_rows(ev, vals, ctx):
    v = vals[0]
    if is_arr(v):
        return arr_shape(v)[0]
    return 1


def f_columns(ev, vals, ctx):
    v = vals[0]
    if is_arr(v):
        return arr_shape(v)[1]
    return 1


def f_index(ev, vals, ctx):
    arr = vals[0]
    if not is_arr(arr):
        arr = [[arr]]
    nr, nc = arr_shape(arr)
    row = vals[1] if len(vals) > 1 else None
    col = vals[2] if len(vals) > 2 else None
    if is_arr(row):
        return [[f_index(ev, [arr, r] + ([col] if col is not None else []), ctx)
                 for r in row_] for row_ in row]
    row = 0 if row is None else (to_num(row) if not isinstance(row, Err) else row)
    if isinstance(row, Err):
        return row
    row = int(row)
    if col is not None:
        col = to_num(col)
        if isinstance(col, Err):
            return col
        col = int(col)
    if nc == 1 and col is None:
        col = 1
    if nr == 1 and col is None and row > 1:
        # single-row vector indexed along its length
        row, col = 1, row
    if row == 0 and col and col >= 1:
        if col > nc:
            return E_REF
        return [[arr[i][col - 1]] for i in range(nr)]
    if col == 0 and row >= 1:
        if row > nr:
            return E_REF
        return [list(arr[row - 1])]
    if row < 1 or row > nr or col < 1 or col > nc:
        return E_REF
    return arr[row - 1][col - 1]


def _match_exact(lookup, vec, wildcard=True):
    for i, x in enumerate(vec):
        if isinstance(x, Err):
            continue
        if isinstance(lookup, str) and wildcard and ("*" in lookup or "?" in lookup):
            if isinstance(x, str) and _wildcard_re(lookup).match(x):
                return i + 1
        elif _eq(x, lookup):
            return i + 1
    return E_NA


def _match_approx(lookup, vec, ascending=True):
    best = E_NA
    for i, x in enumerate(vec):
        if isinstance(x, Err) or x is None:
            continue
        if type(x) is not type(lookup) and not (
                isinstance(x, (int, float)) and isinstance(lookup, (int, float))):
            # skip mismatched types (Excel ordering rules)
            if isinstance(lookup, (int, float)) != isinstance(x, (int, float)):
                continue
        ka, kb = _sort_key(x), _sort_key(lookup)
        c = (ka > kb) - (ka < kb)
        if ascending:
            if c <= 0:
                best = i + 1
            else:
                break
        else:
            if c >= 0:
                best = i + 1
            else:
                break
    return best


def f_match(ev, vals, ctx):
    lookup, arr = vals[0], vals[1]
    mtype = vals[2] if len(vals) > 2 else 1
    if isinstance(lookup, Err):
        return lookup
    if isinstance(arr, Err):
        return arr
    vec = _flat1d(arr)
    if vec is None:
        return E_NA
    mtype = to_num(mtype) if not isinstance(mtype, Err) else mtype
    if isinstance(mtype, Err):
        return mtype
    if mtype == 0:
        return _match_exact(lookup, vec)
    return _match_approx(lookup, vec, ascending=mtype > 0)


def f_xmatch(ev, vals, ctx):
    lookup, arr = vals[0], vals[1]
    mode = vals[2] if len(vals) > 2 else 0
    search = vals[3] if len(vals) > 3 else 1
    if isinstance(lookup, Err):
        return lookup
    vec = _flat1d(arr)
    if vec is None:
        return E_NA
    mode = to_num(mode)
    if isinstance(mode, Err):
        return mode
    search = to_num(search)
    if isinstance(search, Err):
        return search
    reverse = search == -1
    if reverse:
        rvec = list(reversed(vec))
        def flip(r):
            return r if isinstance(r, Err) else len(vec) - r + 1
    else:
        rvec = vec
        def flip(r):
            return r
    if mode == 0:
        return flip(_match_exact(lookup, rvec, wildcard=False))
    if mode == 2:
        return flip(_match_exact(lookup, rvec, wildcard=True))
    if mode == -1:  # exact or next smaller
        r = _match_exact(lookup, rvec, wildcard=False)
        if not isinstance(r, Err):
            return flip(r)
        return _match_approx(lookup, vec, ascending=True)
    if mode == 1:  # exact or next larger
        r = _match_exact(lookup, rvec, wildcard=False)
        if not isinstance(r, Err):
            return flip(r)
        for i, x in enumerate(vec):
            if isinstance(x, (int, float)) and isinstance(lookup, (int, float)) and x >= lookup:
                return i + 1
        return E_NA
    return E_VALUE


def f_lookup(ev, vals, ctx):
    lookup = vals[0]
    vec = _flat1d(vals[1])
    if vec is None:
        return E_NA
    res = _flat1d(vals[2]) if len(vals) > 2 else vec
    pos = _match_approx(lookup, vec, ascending=True)
    if isinstance(pos, Err):
        return pos
    if pos - 1 >= len(res):
        return E_NA
    return res[pos - 1]


def f_countif(ev, vals, ctx):
    rng, crit = vals[0], vals[1]
    if isinstance(rng, Err):
        return rng
    if is_arr(crit):
        return [[f_countif(ev, [rng, c], ctx) for c in flatten(crit)]]
    pred = make_criteria(crit)
    items = flatten(rng) if is_arr(rng) else [rng]
    n = 0
    for x in items:
        r = pred(x)
        if r is True:
            n += 1
    return n


def f_sumif(ev, vals, ctx):
    rng, crit = vals[0], vals[1]
    if isinstance(rng, Err):
        return rng
    sumrng = vals[2] if len(vals) > 2 else rng
    pred = make_criteria(crit)
    items = list(flatten(rng)) if is_arr(rng) else [rng]
    sums = list(flatten(sumrng)) if is_arr(sumrng) else [sumrng]
    total = 0.0
    for i, x in enumerate(items):
        if pred(x) is True and i < len(sums):
            y = sums[i]
            if isinstance(y, (int, float)) and not isinstance(y, bool):
                total += y
    return total


def f_frequency(ev, vals, ctx):
    data = [x for x in flatten(vals[0]) if isinstance(x, (int, float)) and not isinstance(x, bool)]
    bins = [x for x in flatten(vals[1]) if isinstance(x, (int, float)) and not isinstance(x, bool)]
    bins = sorted(bins)
    counts = [0] * (len(bins) + 1)
    for x in data:
        placed = False
        for i, b in enumerate(bins):
            if x <= b:
                counts[i] += 1
                placed = True
                break
        if not placed:
            counts[-1] += 1
    return [[c] for c in counts]


def f_vstack(ev, vals, ctx):
    arrays = [to_array(v) for v in vals]
    width = max(arr_shape(a)[1] for a in arrays)
    out = []
    for a in arrays:
        r, c = arr_shape(a)
        for i in range(r):
            row = [a[i][j] if j < c else E_NA for j in range(width)]
            out.append(row)
    return out


def f_drop(ev, vals, ctx):
    arr = to_array(vals[0])
    nr, nc = arr_shape(arr)
    rows = int(to_num(vals[1])) if len(vals) > 1 and vals[1] is not None else 0
    cols = int(to_num(vals[2])) if len(vals) > 2 and vals[2] is not None else 0
    r0 = rows if rows > 0 else 0
    r1 = nr + rows if rows < 0 else nr
    c0 = cols if cols > 0 else 0
    c1 = nc + cols if cols < 0 else nc
    if r0 >= r1 or c0 >= c1:
        return E_CALC
    return [row[c0:c1] for row in arr[r0:r1]]


def f_take(ev, vals, ctx):
    arr = to_array(vals[0])
    nr, nc = arr_shape(arr)
    rows = int(to_num(vals[1])) if len(vals) > 1 and vals[1] is not None else nr
    cols = int(to_num(vals[2])) if len(vals) > 2 and vals[2] is not None else nc
    rsel = arr[:rows] if rows >= 0 else arr[rows:]
    out = [row[:cols] if cols >= 0 else row[cols:] for row in rsel]
    if not out or not out[0]:
        return E_CALC
    return out


def f_unique(ev, vals, ctx):
    arr = to_array(vals[0])
    seen = set()
    out = []
    for row in arr:
        key = tuple((x.upper() if isinstance(x, str) else x) for x in row)
        if key not in seen:
            seen.add(key)
            out.append(list(row))
    return out


def f_filter(ev, vals, ctx):
    arr = to_array(vals[0])
    inc = to_array(vals[1])
    if_empty = vals[2] if len(vals) > 2 else E_CALC
    nr, nc = arr_shape(arr)
    ir, ic = arr_shape(inc)
    # Excel propagates errors found in the include mask
    for x in flatten(inc):
        if isinstance(x, Err):
            return x
    if ic == nr and ir == 1 and nc != nr:
        # filter columns
        keep = [j for j in range(min(nc, ic)) if to_bool(inc[0][j]) is True]
        if not keep:
            return if_empty
        return [[row[j] for j in keep] for row in arr]
    keep = [i for i in range(min(nr, ir)) if to_bool(inc[i][0]) is True]
    if not keep:
        return if_empty
    return [arr[i] for i in keep]


def f_today(ev, vals, ctx):
    if ev.today is None:
        raise RuntimeError("today not injected")
    return ev.today


def f_stockhistory(ev, vals, ctx):
    if ev.stock_provider is None:
        return Err("#NO_PROVIDER")
    return ev.stock_provider(vals)


def f_row_special(ev, args, ctx):
    if not args or (args[0][0] == "empty"):
        return ctx.cell[0] if ctx.cell else E_VALUE
    n = args[0]
    if n[0] == "ref":
        return n[4]
    if n[0] == "range":
        a, b = n[1], n[2]
        if a[0] == "ref" and b[0] == "ref":
            return [[r] for r in range(a[4], b[4] + 1)]
    return E_VALUE


def f_column_special(ev, args, ctx):
    if not args or (args[0][0] == "empty"):
        return ctx.cell[1] if ctx.cell else E_VALUE
    n = args[0]
    if n[0] == "ref":
        return column_index_from_string(n[3])
    return E_VALUE


# -------------------------------------------------------- lazy specials ----

def sp_if(ev, args, ctx):
    cond = ev.eval_node(args[0], ctx)
    if is_arr(cond) and arr_shape(cond) == (1, 1):
        cond = cond[0][0]
    if is_arr(cond):
        tval = ev.eval_node(args[1], ctx) if len(args) > 1 else True
        fval = ev.eval_node(args[2], ctx) if len(args) > 2 else False
        if is_arr(tval) and arr_shape(tval) == (1, 1):
            tval = tval[0][0]
        if is_arr(fval) and arr_shape(fval) == (1, 1):
            fval = fval[0][0]
        nr, nc = arr_shape(cond)
        out = []
        for i in range(nr):
            row = []
            for j in range(nc):
                c = to_bool(cond[i][j])
                if isinstance(c, Err):
                    row.append(c)
                    continue
                v = tval if c else fval
                if is_arr(v):
                    vr, vc = arr_shape(v)
                    row.append(v[i][j] if i < vr and j < vc else E_NA)
                else:
                    row.append(v)
            out.append(row)
        return out
    c = to_bool(cond)
    if isinstance(c, Err):
        return c
    if c:
        return ev.eval_node(args[1], ctx) if len(args) > 1 else True
    return ev.eval_node(args[2], ctx) if len(args) > 2 else False


def sp_ifs(ev, args, ctx):
    i = 0
    while i + 1 < len(args):
        cond = ev.eval_node(args[i], ctx)
        c = to_bool(first(cond) if is_arr(cond) else cond)
        if isinstance(c, Err):
            return c
        if c:
            return ev.eval_node(args[i + 1], ctx)
        i += 2
    return E_NA


def sp_iferror(ev, args, ctx):
    v = ev.eval_node(args[0], ctx)
    if is_arr(v):
        if any(is_err(x) for x in flatten(v)):
            alt = ev.eval_node(args[1], ctx) if len(args) > 1 else None
            nr, nc = arr_shape(v)
            out = []
            for i in range(nr):
                row = []
                for j in range(nc):
                    if is_err(v[i][j]):
                        if is_arr(alt):
                            ar, ac = arr_shape(alt)
                            row.append(alt[i][j] if i < ar and j < ac else E_NA)
                        else:
                            row.append(alt)
                    else:
                        row.append(v[i][j])
                out.append(row)
            return out
        return v
    if is_err(v):
        return ev.eval_node(args[1], ctx) if len(args) > 1 else None
    return v


def sp_ifna(ev, args, ctx):
    v = ev.eval_node(args[0], ctx)
    if is_err(v) and v.code == "#N/A":
        return ev.eval_node(args[1], ctx) if len(args) > 1 else None
    return v


def sp_let(ev, args, ctx):
    env = dict(ctx.env)
    i = 0
    while i + 1 < len(args) - 1:
        name_node = args[i]
        if name_node[0] != "name":
            return E_VALUE
        val = ev.eval_node(args[i + 1], Ctx(ev, env, ctx.cell))
        env[name_node[1]] = val
        i += 2
    return ev.eval_node(args[-1], Ctx(ev, env, ctx.cell))


def sp_lambda(ev, args, ctx):
    if len(args) < 2:
        return E_VALUE
    params = []
    for a in args[:-1]:
        if a[0] != "name":
            return E_VALUE
        params.append(a[1])
    return XLLambda(params, args[-1], ctx.env)


def apply_lambda(lam, arg_vals, ev, cell):
    env = dict(lam.env)
    for p, v in zip(lam.params, arg_vals):
        env[p] = v
    return ev.eval_node(lam.body, Ctx(ev, env, cell))


def sp_scan(ev, args, ctx):
    init = ev.eval_node(args[0], ctx)
    arr = ev.eval_node(args[1], ctx)
    lam = ev.eval_node(args[2], ctx)
    if isinstance(arr, Err):
        return arr
    if not isinstance(lam, XLLambda):
        return E_VALUE
    items = list(flatten(arr)) if is_arr(arr) else [arr]
    nr, nc = arr_shape(arr) if is_arr(arr) else (1, 1)
    acc = init
    flat_out = []
    for x in items:
        acc = apply_lambda(lam, [acc, x], ev, ctx.cell)
        flat_out.append(acc)
    out = []
    k = 0
    for i in range(nr):
        row = []
        for j in range(nc):
            row.append(flat_out[k])
            k += 1
        out.append(row)
    return out


def sp_iserr_special(ev, args, ctx):
    return None


FUNCS = {
    "SUM": f_sum, "AVERAGE": f_average, "MIN": f_min, "MAX": f_max,
    "MEDIAN": f_median, "STDEV.P": f_stdevp, "STDEV": f_stdev, "SQRT": f_sqrt,
    "ABS": f_abs, "AND": f_and, "OR": f_or,
    "ISNUMBER": f_isnumber, "ISERROR": f_iserror, "ISODD": f_isodd,
    "ISBLANK": f_isblank, "LEN": f_len, "EXACT": f_exact, "MID": f_mid,
    "FIND": f_find, "SEARCH": f_search, "CONCAT": f_concat,
    "TEXTJOIN": f_textjoin, "TEXTSPLIT": f_textsplit,
    "ROWS": f_rows, "COLUMNS": f_columns, "INDEX": f_index,
    "MATCH": f_match, "XMATCH": f_xmatch, "LOOKUP": f_lookup,
    "COUNTIF": f_countif, "SUMIF": f_sumif, "FREQUENCY": f_frequency,
    "VSTACK": f_vstack, "DROP": f_drop, "TAKE": f_take, "UNIQUE": f_unique,
    "FILTER": f_filter, "TODAY": f_today, "STOCKHISTORY": f_stockhistory,
}

SPECIAL = {
    "IF": sp_if, "IFS": sp_ifs, "IFERROR": sp_iferror, "IFNA": sp_ifna,
    "LET": sp_let, "LAMBDA": sp_lambda, "SCAN": sp_scan,
    "ROW": f_row_special, "COLUMN": f_column_special,
}
