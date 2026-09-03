"""Run: python -m src.ticker_lookback_run --tickers TEM,ELF,AAPL"""
from __future__ import annotations

import argparse
import html
import json
import os
from datetime import datetime

from . import gainer_asof as ga
from . import lookback_action as act
from . import ticker_lookback as tl
from . import ticker_lookback_cli as scan
from . import ticker_lookback_setups as setups

CAMERA_COLS = tl.BOX_COLS + (("yday", "yΔ"),)
DOMAIN_COLS = ga.DOMAIN_COLS


def _icon(kind):
    kind = {"green": "good", "red": "bad", "yellow": "neutral"}.get(
        str(kind or "").lower(), kind)
    return tl.BOX_ICON.get(kind, tl.BOX_ICON["missing"])


def _price_tones(pc):
    pc = pc or {}
    return {k: tl.price_tone(pc.get(k)) for k in ("1d", "3d", "1w")}


def _attach_day_extras(card, ticker, sess, sessions):
    fv = (sess.get("finviz") or {}).get(ticker)
    card["price_changes"] = tl.forward_price_changes(
        ticker, sess["date"], sessions=sessions, current_finviz=fv)
    card["forward_returns"] = tl.forward_returns(
        ticker, sess["date"], sessions=sessions, current_finviz=fv)
    card["price_tones"] = _price_tones(card["price_changes"])
    card["session_bar"] = tl.session_bar(ticker, sess["date"])
    card["horizon_dates"] = tl.horizon_dates(sess["date"])
    return card


def _paint_cache(sessions):
    cache = {}
    for sess in sessions or []:
        date = sess["date"]
        if date in cache:
            continue
        ctx = ga.load_day_context(date)
        ctx["buy_today"] = ga.same_day_buy_set(date)
        ctx["sell_today"] = ga.same_day_sell_set(date)
        ctx["era_skip"] = ga._era_skip(date)
        cache[date] = ctx
    return cache


def _asof_decision(card):
    lane = card.get("lane")
    state = str(card.get("market_state") or "").upper() or "UNKNOWN"
    if not card.get("lattice_live"):
        return "—"
    if not lane:
        return "—"
    if lane == "probable":
        return f"BUY PROBABLE — most-probable long on {state} (size ×0.25)"
    if lane == "blocked":
        return f"BLOCK BUY — market={state}"
    if lane == "catalyst_exception":
        return f"BUY CATALYST_EXCEPTION — market={state}"
    return f"BUY {str(lane).upper()} — market={state}"


def _decision_text(card):
    action = card.get("action") or {}
    if action.get("bull_eligible") and action.get("bull_decision"):
        return action["bull_decision"]
    if action.get("bear_eligible") and action.get("bear_decision"):
        return action["bear_decision"]
    if action.get("bull_decision"):
        return action["bull_decision"]
    if action.get("bear_decision"):
        return action["bear_decision"]
    return _asof_decision(card)


def _context_text(card):
    action = card.get("action") or {}
    bits = []
    summary = str(action.get("company_summary") or "").strip()
    if summary:
        bits.append(summary)
    group = action.get("group_label") or card.get("industry") or ""
    d1, w1, rel = action.get("child_d1"), action.get("child_w1"), action.get("child_residual")
    if group and any(v is not None for v in (d1, w1, rel)):
        bits.append(
            f"{group} "
            f"({0.0 if d1 is None else d1:+.1f}% d1 / "
            f"{0.0 if w1 is None else w1:+.1f}% 1w / "
            f"{0.0 if rel is None else rel:+.1f}% rel)"
        )
    elif card.get("sector"):
        bits.append(str(card["sector"]))
    return "; ".join(bits)


def _decision_cell(card):
    ctx = _context_text(card)
    why = card.get("decision") or "—"
    if ctx and why and why != "—":
        return f"{ctx} — {why}"
    return ctx or why


def _hall_text(day):
    label = day.get("lane_label")
    if label is None:
        label = ga.lane_label(day.get("lane"))
    state = str(day.get("market_state") or "").lower()
    if state == "hard_red":
        if label and label != ga.GREY:
            return f"{label} · HARD_RED"
        return "HARD_RED"
    return label or "—"


def _paint_day(card, sess, prior_sess, ctx):
    t = card["ticker"]
    painted = ga.color_name(
        sess,
        {
            "ticker": t,
            "sector": card.get("sector") or "",
            "industry": card.get("industry") or "",
            "size": card.get("size") or "",
            "decision_lane": (ctx.get("lanes") or {}).get(t),
        },
        buy_today=ctx.get("buy_today") or set(),
        sell_today=ctx.get("sell_today") or set(),
        market_tone=ctx.get("market_tone"),
        market_state=ctx.get("market_state"),
        book_domains=ctx.get("domains"),
        book_lanes=ctx.get("lanes"),
        book_marks=ctx.get("marks"),
        book_opp=ctx.get("opp"),
        era_skip=ctx.get("era_skip"),
        lattice_live=ctx.get("lattice_live"),
        prior_sess=prior_sess,
        card=card,
        prev_boxes=(prior_sess or {}).get("_prev_boxes"),
    )
    card["boxes"] = dict(card.get("boxes") or {})
    card["boxes"]["yday"] = (painted.get("boxes") or {}).get("yday") or "missing"
    vintage = dict(card.get("factor_vintage") or {})
    for key in ("yday", "domains"):
        if (painted.get("factor_vintage") or {}).get(key):
            vintage[key] = painted["factor_vintage"][key]
    card["factor_vintage"] = vintage
    for key in (
        "domains", "labeled", "labeled_domains", "lane", "lane_label",
        "marks", "marks_cell", "mid_opp", "yday_change",
        "overnight_buy", "overnight_sell", "on_1d_buy", "on_1d_sell",
        "lattice_live",
    ):
        card[key] = painted.get(key)
    card["market_state"] = ctx.get("market_state")
    card["market_tone"] = ctx.get("market_tone")
    card["action"] = (ctx.get("actions") or {}).get(t) or {}
    card["decision"] = _decision_text(card)
    return card


def scan_ticker(ticker, sessions=None, idx=None, paint_ctx=None):
    idx = idx or tl.build_index()
    t = tl._tick(ticker)
    days, recommended, green_days = [], [], []
    sessions = sessions if sessions is not None else idx["sessions"]
    paint_ctx = paint_ctx if paint_ctx is not None else _paint_cache(sessions)
    prev_card = None
    for sess in sessions:
        card = scan._scan_session(sess, t)
        if card is None:
            card = {
                "date": sess["date"], "ticker": t, "class": "no_data",
                "sources": [], "boxes": {k: "missing" for k, _ in tl.BOX_COLS},
                "artifacts_that_day": sess["has"],
            }
        _attach_day_extras(card, t, sess, sessions)
        prior = sess.get("prior")
        if prev_card is not None:
            prior = dict(prior or {})
            prior["_prev_boxes"] = prev_card.get("boxes")
        _paint_day(card, sess, prior, paint_ctx.get(sess["date"]) or {})
        days.append(card)
        prev_card = card
        if card.get("buy_ranks"):
            recommended.append({
                "date": card["date"],
                "horizons": list(card["buy_ranks"].keys()),
                "ranks": card["buy_ranks"],
            })
        if card.get("independent_green", {}).get("green") or card.get("in_green_buy"):
            green_days.append(card["date"])
    tl.annotate_signal_improved(days)
    hits = [d for d in days if d.get("class") != "no_data"]
    return {
        "ticker": t, "n_sessions": len(sessions), "n_with_print": len(hits),
        "recommended_days": recommended, "green_days": green_days,
        "paper": idx["paper"].get(t) or [], "days": days,
    }


def scan_tickers(tickers, from_date=None, to_date=None):
    names = [tl._tick(t) for t in tickers if tl._tick(t)]
    idx = tl.build_index()
    sessions = [
        s for s in idx["sessions"]
        if (not from_date or s["date"] >= from_date)
        and (not to_date or s["date"] <= to_date)
    ]
    paint_ctx = _paint_cache(sessions)
    return {
        "generated_at": datetime.now(tl.ET).isoformat(),
        "asof": "09:30_et",
        "method": "stock_book_readiness",
        "from_date": from_date, "to_date": to_date,
        "sessions": [
            {"date": s["date"], "has": s["has"], "n_book": s["n_book"],
             "n_join": s["n_join"], "n_finviz": s["n_finviz"],
             "n_ab": s["n_ab"], "n_peer": s["n_peer"]}
            for s in sessions
        ],
        "names": [
            scan_ticker(t, sessions=sessions, idx=idx, paint_ctx=paint_ctx)
            for t in names
        ],
    }


def _fmt_price(pc, key):
    v = (pc or {}).get(key)
    if v is None:
        return "—"
    return f"{v:+.2f}%" if key != "price" else f"${v:,.2f}"


def _fmt_price_md(pc, tones, key):
    text = _fmt_price(pc, key)
    if key == "price" or text == "—":
        return text
    return f"{_icon((tones or {}).get(key) or tl.price_tone((pc or {}).get(key)))} {text}"


def _mark_flags(day):
    packed = day.get("marks")
    if isinstance(packed, dict) and any(
        packed.get(k) is not None for k in ("blue", "alarm", "white")
    ):
        return (
            bool(packed.get("blue")),
            bool(packed.get("alarm")),
            bool(packed.get("white")),
        )
    return (
        bool(day.get("signal_improved")),
        bool(day.get("signal_alarm")),
        bool(day.get("zero_red")),
    )


def _date_marks(day):
    blue, alarm, white = _mark_flags(day)
    return "".join(
        bit for bit, on in (("🔵", blue), ("🚨", alarm), ("⚪", white)) if on
    )


def _date_label(day):
    marks = _date_marks(day)
    stamp = act.session_stamp(day.get("date"), act.OPEN_CLOCK)
    return f"{marks} {stamp}".strip() if marks else stamp


def _date_classes(day):
    cls = []
    blue, alarm, white = _mark_flags(day)
    if blue:
        cls.append("better")
    if alarm:
        cls.append("alarm")
    if white:
        cls.append("clean")
    tone = (day.get("region") or {}).get("tone")
    if tone == "good":
        cls.append("reg-good")
    elif tone == "bad":
        cls.append("reg-bad")
    lane = day.get("lane")
    if lane == "probable":
        cls.append("probable")
    elif lane == "blocked":
        cls.append("blocked")
    if str(day.get("market_state") or "").lower() == "hard_red":
        cls.append("hard-red")
    return " ".join(cls)


def _condition(day):
    cond = day.get("condition")
    if not cond:
        cond = tl.general_condition(day.get("boxes") or {})
    return cond


def _condition_text(cond):
    if not cond or cond.get("tone") == "missing" or not cond.get("n"):
        return "—"
    return f"{cond['good']}/{cond['neutral']}/{cond['bad']}"


def _condition_md(cond):
    text = _condition_text(cond)
    if text == "—":
        return text
    return f"{_icon(cond.get('tone'))} {text}"


def _region(day):
    reg = day.get("region")
    if not reg:
        reg = tl.color_region(day.get("boxes") or {})
    return reg


def _region_text(reg):
    tone = (reg or {}).get("tone")
    if tone in (None, "missing", "thin"):
        return "—"
    g, r = (reg or {}).get("good", 0), (reg or {}).get("bad", 0)
    return f"{g}-{r}"


def _region_md(reg):
    text = _region_text(reg)
    if text == "—":
        return text
    return f"{_icon((reg or {}).get('tone'))} {text}"


def _fmt_mid_opp(value):
    if value is None:
        return "—"
    try:
        return f"{float(value):+.2f}"
    except (TypeError, ValueError):
        return "—"


def _sheet_headers():
    return (
        ["Date", "Price", "Open", "o→c", "+1d", "+3d", "+1w", "Action", "Cond", "Hall pass", "mid_opp", "Setups"]
        + [label for _, label in CAMERA_COLS]
        + [label for _, label in DOMAIN_COLS]
        + ["Decision"]
    )


def _action_text(day):
    return day.get("action_label") or act.format_action(
        day.get("action_call"), day.get("date"))


def _session_bar(day):
    bar = day.get("session_bar")
    if bar:
        return bar
    ticker = day.get("ticker")
    date = day.get("date")
    if ticker and date:
        return tl.session_bar(ticker, date)
    return {}


def _horizon_dates(day):
    hz = day.get("horizon_dates")
    if hz:
        return hz
    date = day.get("date")
    return tl.horizon_dates(date) if date else {}


def _price_cell(day):
    bar = _session_bar(day)
    px = bar.get("close")
    if px is None:
        px = (day.get("price_changes") or {}).get("price")
    return act.format_price(px, day.get("date"), act.CLOSE_CLOCK)


def _open_cell(day):
    return act.format_price(_session_bar(day).get("open"), day.get("date"), act.OPEN_CLOCK)


def _open_close_cell(day):
    return act.format_open_close(_session_bar(day).get("close_open_pct"), day.get("date"))


def _fwd_cell(day, key):
    pc = day.get("price_changes") or {}
    when = _horizon_dates(day).get(key)
    return act.format_ret(pc.get(key), when, act.CLOSE_CLOCK)


def render_md(payload):
    setups.ensure_setups(payload)
    act.ensure_actions(payload)
    L = ["# Ticker lookback", "", f"_Generated {payload['generated_at']}_",
         "", "_Same method as Stock Book readiness._ Cameras (12 + yΔ) are "
         "the 09:30 ET packet: last completed tape (walk back if a session "
         "file is missing) + that morning's pre-open. Same-day stock book "
         "and post-close Finviz do not color cameras. Domain lights, hall "
        "pass, and BUY PROBABLE / HARD_RED copy come from that day's book "
        "when it exists, else the as-of coaches. **Action** is the "
        "authoritative BUY / SELL / NO BUY / HOLD from those 09:30 "
        "cameras, featured mine setups, and hall pass. "
        "**Action is known at that date 09:30 ET** — the regular open, "
        "before anyone knows the name will close as a gainer. It is not "
        "an end-of-day call to trade the next morning. Price is the "
        "16:00 ET close; Open is 09:30 ET; o→c is open→close the same "
        "session; +1d / +3d / +1w are later 16:00 ET closes._", ""]
    if payload.get("random"):
        L += [f"_Random {len(payload['names'])} names, "
              f"mcap > $100M, avg vol > 500K, ATR% ≥ {tl.MIN_ATR_PCT}_", ""]
    L += [setups.render_setup_markdown(payload, include_dates=False), ""]
    L += [
        f"_Cameras {ga._legend()}_",
        f"_Coaches {ga._domain_legend()}_",
        "_Hall pass = standard / group leader / catalyst / catalyst exception "
        "/ probable / blocked — or grey before the lattice (2026-08-31). "
        "HARD_RED may still print BUY PROBABLE (size ×0.25)._",
        "_🔵 = vs prior session: no cell worse and at least one better, "
        "or factor points jumped by ≥3 (red=1, yellow=2, green=3)_",
        "_🚨 = purely worse vs prior session (no cell better, at least one worse)_",
        "_⚪ = no red factor cells that day_",
        "_Cond = G/Y/R tally; green or red when that color is the majority. "
        "Setups overlay the color chart on the date they printed — bare "
        "🔵 / 🚨 / ⚪ and 🔵-on-red (`turn`) did not replicate market-wide._",
        "",
    ]
    heads = _sheet_headers()
    bars = "|".join(["---"] * len(heads))
    for rec in payload["names"]:
        L += [f"## {rec['ticker']}", "",
              "| " + " | ".join(heads) + " |",
              f"|{bars}|"]
        for d in rec["days"]:
            pc = d.get("price_changes") or {}
            tones = d.get("price_tones") or _price_tones(pc)
            boxes = d.get("boxes") or {}
            domains = d.get("domains") or {}
            cams = " | ".join(
                _icon(boxes.get(k, "missing")) for k, _ in CAMERA_COLS)
            coaches = " | ".join(
                _icon(domains.get(k, "missing")) for k, _ in DOMAIN_COLS)
            date = _date_label(d)
            why = str(_decision_cell(d)).replace("|", "/")
            L.append(
                f"| {date} | {_price_cell(d)} | {_open_cell(d)} | {_open_close_cell(d)} | "
                f"{_fwd_cell(d, '1d')} | "
                f"{_fwd_cell(d, '3d')} | "
                f"{_fwd_cell(d, '1w')} | "
                f"**{_action_text(d)}** | "
                f"{_condition_md(_condition(d))} | "
                f"{_hall_text(d)} | {_fmt_mid_opp(d.get('mid_opp'))} | "
                f"{setups.setup_labels(d) or '—'} | {cams} | {coaches} | {why} |"
            )
        L.append("")
    return "\n".join(L) + "\n"


def _slug(tickers, random_pick=False):
    names = [tl._tick(t) for t in tickers if tl._tick(t)]
    if random_pick:
        head = "-".join(t.lower() for t in names[:4])
        n = len(names)
        if head:
            return f"random-{n}-{head}"
        return f"random-{n}"
    return "-".join(t.lower() for t in names)


def render_html(payload):
    setups.ensure_setups(payload)
    act.ensure_actions(payload)
    sections = []
    for rec in payload["names"]:
        rows = []
        for day in rec["days"]:
            pc = day.get("price_changes") or {}
            tones = day.get("price_tones") or _price_tones(pc)
            lit = setups.box_highlights(day)
            cells = "".join(
                f'<td class="{html.escape((day.get("boxes") or {}).get(k, "missing"))}'
                f'{" setup-hit setup-" + html.escape(lit[k]) if k in lit else ""}">'
                f'{_icon((day.get("boxes") or {}).get(k, "missing"))}</td>'
                for k, _ in CAMERA_COLS
            )
            date_cls = _date_classes(day)
            cond = _condition(day)
            oc_tone = tl.price_tone(_session_bar(day).get("close_open_pct"))
            price_tds = "".join(
                f'<td class="{html.escape(tones.get(key, "missing"))}">'
                f'{html.escape(_fwd_cell(day, key))}</td>'
                for key in ("1d", "3d", "1w")
            )
            chips = setups.setup_chips_html(day)
            row_cls = setups.row_setup_class(day)
            hall = _hall_text(day)
            hall_cls = html.escape(str(day.get("lane") or "missing"))
            domains = day.get("domains") or {}
            domain_tds = "".join(
                f'<td class="{html.escape(domains.get(k, "missing"))}">'
                f'{_icon(domains.get(k, "missing"))}</td>'
                for k, _ in DOMAIN_COLS
            )
            why = html.escape(_decision_cell(day))
            action = _action_text(day)
            action_tone = act.action_tone(day.get("action_call") or "")
            rows.append(
                f'<tr class="{html.escape(row_cls)}">'
                f'<th class="{html.escape(date_cls)}">'
                f'{html.escape(_date_label(day))}</th>'
                f"<td>{html.escape(_price_cell(day))}</td>"
                f"<td>{html.escape(_open_cell(day))}</td>"
                f'<td class="{html.escape(oc_tone)}">'
                f'{html.escape(_open_close_cell(day))}</td>'
                f"{price_tds}"
                f'<td class="action {html.escape(action_tone)}">'
                f'{html.escape(action)}</td>'
                f'<td class="{html.escape(cond.get("tone", "missing"))}">'
                f'{html.escape(_condition_text(cond))}</td>'
                f'<td class="hall {hall_cls}">{html.escape(hall)}</td>'
                f'<td class="mid-opp">{html.escape(_fmt_mid_opp(day.get("mid_opp")))}</td>'
                f'<td class="setups">{chips or "—"}</td>{cells}'
                f'{domain_tds}'
                f'<td class="decision">{why}</td></tr>'
            )
        factor_headers = "".join(
            f"<th>{html.escape(label)}</th>" for _, label in CAMERA_COLS)
        domain_headers = "".join(
            f"<th>{html.escape(label)}</th>" for _, label in DOMAIN_COLS)
        sections.append(f"""
<section class="ticker" id="{html.escape(rec['ticker'])}">
 <h2>{html.escape(rec['ticker'])}</h2>
 <div class="sheet"><table>
 <thead><tr><th>Date</th><th>Price</th><th>Open</th><th>o→c</th><th>+1d</th><th>+3d</th><th>+1w</th><th>Action</th><th>Cond</th><th>Hall pass</th><th>mid_opp</th><th>Setups</th>{factor_headers}{domain_headers}<th>Decision</th></tr></thead>
 <tbody>{''.join(rows)}</tbody></table></div>
</section>""")
    nav = '<a href="#setups">Setups</a>' + "".join(
        f'<a href="#{html.escape(r["ticker"])}">{html.escape(r["ticker"])}</a>'
        for r in payload["names"]
    )
    picked = ", ".join(html.escape(r["ticker"]) for r in payload["names"])
    random_note = (
        f'<p class="muted">Random draw: {picked} · mcap &gt; $100M · avg vol &gt; 500K</p>'
        if payload.get("random") else ""
    )
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Ticker lookback</title>
<style>
:root{{--bg:#0b1020;--card:#131b31;--line:#2b3552;--text:#edf2ff;--muted:#9cabc9}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--text);font:15px/1.45 system-ui}}
main{{max-width:1000px;margin:auto;padding:16px}}h1,h2,h3,h4{{margin:.35em 0}}
nav{{display:flex;gap:8px;overflow:auto;position:sticky;top:0;background:#0b1020ee;padding:10px 0;z-index:2}}
nav a,.class{{padding:8px 12px;border:1px solid var(--line);border-radius:999px;color:var(--text);text-decoration:none;white-space:nowrap}}
.sheet{{overflow-x:auto;border:1px solid var(--line);border-radius:12px;margin-bottom:22px}}
table{{border-collapse:separate;border-spacing:0;min-width:1680px;width:100%;background:var(--card)}}
section.setups table{{min-width:640px}}
th,td{{padding:10px 9px;text-align:center;border-bottom:1px solid var(--line);white-space:nowrap}}
thead th{{position:sticky;top:0;background:#17213a}}tbody th{{position:sticky;left:0;background:#17213a;text-align:left}}
td.good{{background:#123d2c}}td.bad{{background:#4b2028}}td.neutral{{background:#473e1d}}td.missing{{background:#23283a}}
td.setup-hit{{outline:2px solid #eab308;outline-offset:-2px}}
td.setup-hit.setup-fade{{outline-color:#fb923c}}
tr.has-setup td.setups{{box-shadow:inset 3px 0 0 #eab308}}
tr.setup-fade td.setups{{box-shadow:inset 3px 0 0 #fb923c}}
tbody th.better{{background:#1d4ed8;color:#edf2ff}}
tbody th.alarm{{box-shadow:inset 3px 0 0 #f97316}}
tbody th.alarm:not(.better):not(.clean){{background:#4b2028;color:#edf2ff}}
tbody th.clean{{box-shadow:inset 3px 0 0 #f8fafc}}
tbody th.clean:not(.better){{background:#e8eef7;color:#0b1020}}
tbody th.reg-good{{box-shadow:inset 0 -3px 0 #22c55e}}
tbody th.reg-bad{{box-shadow:inset 0 -3px 0 #ef4444}}
td.setups{{text-align:left;white-space:normal;max-width:200px;font-size:12px}}
td.action{{font-weight:700;font-size:12px}}
td.hall{{font-size:12px;white-space:normal;max-width:140px}}
td.hall.probable{{background:#1e3a5f}}
td.hall.blocked{{background:#4b2028}}
td.decision{{text-align:left;white-space:normal;max-width:360px;font-size:12px}}
tbody th.hard-red{{box-shadow:inset 0 -3px 0 #f59e0b}}
.setup-chip{{display:inline-block;margin:1px 2px;padding:1px 7px;border-radius:999px;font-size:11px;white-space:nowrap}}
.setup-chip.good{{background:#123d2c}}
.setup-chip.bad{{background:#4b2028}}
.muted{{color:var(--muted)}}
@media(max-width:600px){{main{{padding:8px}}th,td{{padding:9px 7px;font-size:13px}}}}
</style></head><body><main>
<h1>Ticker lookback</h1>
<p>Same method as Stock Book readiness. Cameras = knowable by 09:30 ET (last completed tape + pre-open packet). <b>Action is that date 09:30 ET</b> — known at the open, before anyone knows the name will close as a gainer. Not an end-of-day call for the next morning. Price = 16:00 ET close. Open = 09:30 ET. o→c = same-session open→close. +1d / +3d / +1w = later 16:00 ET closes.</p>
<p>🟢 up / positive · 🟡 flat · 🔴 down / negative · ⬛ missing · ⬜ no as-of · 🔵 improved or +≥3 pts · 🚨 purely worse · ⚪ no red · Cond = G/Y/R majority · Hall pass = standard / group leader / catalyst / probable / blocked · Action = BUY / SELL / NO BUY / HOLD</p>
<p class="muted">Two scoreboards: 12 cameras + yΔ, then 6 coaches (mkt · par · chd · co · set · flw). HARD_RED may still print BUY PROBABLE (size ×0.25). Featured setups overlay the camera sheet: gold ring on the boxes that fired.</p>
{random_note}{setups.render_setup_html(payload)}<nav>{nav}</nav>{''.join(sections)}
</main></body></html>"""


def _write_setups_sheet(wb, payload, fills):
    from openpyxl.styles import Alignment, Font, PatternFill

    setups.ensure_setups(payload)
    window = payload.get("setup_window") or setups.mine_window()
    ws = wb.create_sheet("Setups", 0)
    ws["A1"] = "Setups that paid market-wide"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        f"Mine window: {window.get('from_date')} → {window.get('to_date')} · "
        f"{window.get('n_tickers')} liquid names · {window.get('n_printed')} printed days"
    )
    ws["A3"] = (
        "Edge is excess vs the same-day universe median, minus the +0.27 sample-mean tilt. "
        "Bare 🔵 / 🚨 / ⚪ and 🔵-on-red (turn) did not replicate."
    )
    headers = [
        "Setup", "Use", "When", "Market n", "1d edge", "3d xs", "1w xs",
        "Mine from", "Mine to", "Hits this run", "This-run +1d",
    ]
    ws.append([])
    ws.append(headers)
    head_row = ws.max_row
    for cell in ws[head_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    run_by = {s["id"]: s for s in (payload.get("setup_this_run") or [])}
    for s in payload.get("setup_book") or []:
        r = run_by.get(s["id"]) or {}
        ws.append([
            s.get("label"), s.get("verdict"), s.get("when"), s.get("n"),
            s.get("edge_1d"), s.get("edge_3d"), s.get("edge_1w"),
            s.get("mine_from"), s.get("mine_to"),
            r.get("hits_this_run") or 0, r.get("this_run_mean_1d"),
        ])
        row = ws.max_row
        tone = "good" if s.get("verdict") == "long" else "bad"
        ws.cell(row, 2).fill = fills.get(tone, fills["missing"])
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        for col in (5, 6, 7, 11):
            ws.cell(row, col).number_format = '0.00'
    ws.append([])
    ws.append(["Dates these setups printed (this run)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    hit_headers = [
        "Date", "Ticker", "Setup", "Use", "This +1d", "This +3d", "This +1w",
        "Market 1d edge", "Market n", "Mine from", "Mine to",
    ]
    ws.append(hit_headers)
    hit_head = ws.max_row
    for cell in ws[hit_head]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for h in payload.get("setup_hits") or []:
        ws.append([
            h.get("date"), h.get("ticker"), h.get("label"), h.get("verdict"),
            h.get("this_1d"), h.get("this_3d"), h.get("this_1w"),
            h.get("edge_1d"), h.get("n"), h.get("mine_from"), h.get("mine_to"),
        ])
        row = ws.max_row
        tone = "good" if h.get("verdict") == "long" else "bad"
        ws.cell(row, 4).fill = fills.get(tone, fills["missing"])
        for col in (5, 6, 7, 8):
            ws.cell(row, col).number_format = '0.00'
    widths = {"A": 36, "B": 12, "C": 56, "D": 12, "E": 12, "F": 12,
              "G": 12, "H": 16, "I": 12, "J": 14, "K": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_xlsx(payload, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    setups.ensure_setups(payload)
    act.ensure_actions(payload)
    gold = Border(
        left=Side(style="medium", color="EAB308"),
        right=Side(style="medium", color="EAB308"),
        top=Side(style="medium", color="EAB308"),
        bottom=Side(style="medium", color="EAB308"),
    )
    fills = {
        "good": PatternFill("solid", fgColor="63BE7B"),
        "neutral": PatternFill("solid", fgColor="FFEB84"),
        "bad": PatternFill("solid", fgColor="F8696B"),
        "missing": PatternFill("solid", fgColor="808080"),
        "better": PatternFill("solid", fgColor="5B9BD5"),
        "clean": PatternFill("solid", fgColor="FFFFFF"),
    }
    wb = Workbook()
    wb.remove(wb.active)
    _write_setups_sheet(wb, payload, fills)
    headers = _sheet_headers()
    col_of = {name: i + 1 for i, name in enumerate(headers)}
    col_of["1d"] = col_of["+1d"]
    col_of["3d"] = col_of["+3d"]
    col_of["1w"] = col_of["+1w"]
    cam_start = col_of["join"]
    domain_start = col_of["mkt"]
    decision_col = col_of["Decision"]
    action_col = col_of["Action"]
    cond_col = col_of["Cond"]
    hall_col = col_of["Hall pass"]
    opp_col = col_of["mid_opp"]
    setups_col = col_of["Setups"]
    for rec in payload["names"]:
        ws = wb.create_sheet(rec["ticker"][:31])
        ws.freeze_panes = "B2"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for day in rec["days"]:
            pc = day.get("price_changes") or {}
            tones = day.get("price_tones") or _price_tones(pc)
            cond = _condition(day)
            domains = day.get("domains") or {}
            bar = _session_bar(day)
            ws.append([
                _date_label(day),
                _price_cell(day),
                _open_cell(day),
                bar.get("close_open_pct"),
                pc.get("1d"),
                pc.get("3d"), pc.get("1w"), _action_text(day),
                _condition_text(cond),
                _hall_text(day), _fmt_mid_opp(day.get("mid_opp")),
                setups.setup_labels(day),
            ] + [
                _icon((day.get("boxes") or {}).get(k, "missing"))
                for k, _ in CAMERA_COLS
            ] + [
                _icon(domains.get(k, "missing"))
                for k, _ in DOMAIN_COLS
            ] + [_decision_cell(day)])
            row = ws.max_row
            date_cell = ws.cell(row, 1)
            blue, alarm, white = _mark_flags(day)
            if blue:
                date_cell.fill = fills["better"]
                date_cell.font = Font(bold=True, color="FFFFFF")
            elif alarm:
                date_cell.fill = fills["bad"]
                date_cell.font = Font(bold=True, color="FFFFFF")
            elif white:
                date_cell.fill = fills["clean"]
                date_cell.font = Font(bold=True, color="1F4E78")
            oc_col = col_of.get("o→c")
            if oc_col:
                oc_cell = ws.cell(row, oc_col)
                oc_cell.number_format = '0.00"%"'
                oc_cell.fill = fills.get(
                    tl.price_tone(bar.get("close_open_pct")), fills["missing"])
                oc_cell.alignment = Alignment(horizontal="center")
            for key in ("1d", "3d", "1w"):
                cell = ws.cell(row, col_of[key])
                cell.number_format = '0.00"%"'
                cell.fill = fills.get(tones.get(key, "missing"), fills["missing"])
                cell.alignment = Alignment(horizontal="center")
            action_cell = ws.cell(row, action_col)
            action_cell.fill = fills.get(
                act.action_tone(day.get("action_call") or ""), fills["missing"])
            action_cell.alignment = Alignment(horizontal="center")
            cond_cell = ws.cell(row, cond_col)
            cond_cell.fill = fills.get(cond.get("tone", "missing"), fills["missing"])
            cond_cell.alignment = Alignment(horizontal="center")
            ws.cell(row, hall_col).alignment = Alignment(horizontal="center", wrap_text=True)
            ws.cell(row, opp_col).alignment = Alignment(horizontal="center")
            ws.cell(row, setups_col).alignment = Alignment(horizontal="left", wrap_text=True)
            lit = setups.box_highlights(day)
            for offset, (key, _label) in enumerate(CAMERA_COLS, start=cam_start):
                tone = (day.get("boxes") or {}).get(key, "missing")
                cell = ws.cell(row, offset)
                cell.fill = fills.get(tone, fills["missing"])
                cell.alignment = Alignment(horizontal="center")
                if key in lit:
                    cell.border = gold
            for offset, (key, _label) in enumerate(DOMAIN_COLS, start=domain_start):
                tone = domains.get(key, "missing")
                cell = ws.cell(row, offset)
                cell.fill = fills.get(tone, fills["missing"])
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row, decision_col).alignment = Alignment(
                horizontal="left", wrap_text=True)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions[_col_letter(col_of["Open"])].width = 28
        ws.column_dimensions[_col_letter(col_of["o→c"])].width = 28
        ws.column_dimensions[_col_letter(action_col)].width = 28
        ws.column_dimensions[_col_letter(hall_col)].width = 18
        ws.column_dimensions[_col_letter(setups_col)].width = 28
        ws.column_dimensions[_col_letter(decision_col)].width = 56
        wide = {action_col, hall_col, setups_col, decision_col,
                col_of["Open"], col_of["o→c"]}
        for col in range(3, len(headers) + 1):
            if col in wide:
                continue
            ws.column_dimensions[_col_letter(col)].width = 9
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def _col_letter(n):
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def resolve_tickers(raw, random_pick=False, n=tl.RANDOM_N, asof=None, seed=None):
    tokens = [t.strip() for t in (raw or "").split(",") if t.strip()]
    named = [t for t in tokens if t.lower() != "random"]
    want_random = bool(random_pick) or any(t.lower() == "random" for t in tokens)
    if want_random:
        return tl.pick_random_tickers(n=n, asof=asof, seed=seed), True
    return [tl._tick(t) for t in named if tl._tick(t)], False


def _emit_github_env(slug, tickers, random_pick=False):
    path = os.environ.get("GITHUB_ENV")
    if not path:
        return
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(f"SLUG={slug}\n")
        fh.write(f"LOOKBACK_TICKERS={','.join(tickers)}\n")
        fh.write(f"LOOKBACK_RANDOM={'true' if random_pick else 'false'}\n")


def _emit_github_summary(payload):
    path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not path:
        return
    with open(path, "a", encoding="utf-8") as fh:
        fh.write(setups.render_setup_markdown(payload))
        fh.write("\n")


def run(tickers, from_date=None, to_date=None, random_pick=False):
    payload = scan_tickers(tickers, from_date=from_date, to_date=to_date)
    payload["random"] = bool(random_pick)
    setups.attach_setups(payload)
    act.attach_actions(payload)
    tl.BOOK_DIR.mkdir(parents=True, exist_ok=True)
    tl.DAILY.mkdir(parents=True, exist_ok=True)
    tl.SCORE.mkdir(parents=True, exist_ok=True)
    slug = _slug(tickers, random_pick=random_pick)
    if not slug:
        raise SystemExit("no valid tickers")
    js = tl.BOOK_DIR / "ticker_lookback.json"
    js.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    json_dir = tl.BOOK_DIR / "ticker_lookback"
    json_dir.mkdir(parents=True, exist_ok=True)
    (json_dir / f"{slug}.json").write_text(
        json.dumps(payload, indent=2, default=str), encoding="utf-8")
    md = render_md(payload)
    (tl.DAILY / "ticker_lookback.md").write_text(md, encoding="utf-8")
    (tl.SCORE / "TICKER_LOOKBACK.md").write_text(md, encoding="utf-8")
    md_dir = tl.SCORE / "ticker_lookback"
    md_dir.mkdir(parents=True, exist_ok=True)
    (md_dir / f"{slug}.md").write_text(md, encoding="utf-8")
    web_dir = tl.ROOT / "dashboard" / "ticker-lookback"
    web_dir.mkdir(parents=True, exist_ok=True)
    page = render_html(payload)
    (web_dir / f"{slug}.html").write_text(page, encoding="utf-8")
    (web_dir / "index.html").write_text(page, encoding="utf-8")
    xlsx_dir = tl.SCORE / "ticker_lookback"
    xlsx_path = xlsx_dir / f"{slug}.xlsx"
    write_xlsx(payload, xlsx_path)
    _emit_github_env(slug, tickers, random_pick=random_pick)
    _emit_github_summary(payload)
    print(md[:12000])
    print(f"[ticker-lookback] slug {slug}")
    print(f"[ticker-lookback] names {','.join(tickers)}")
    print(f"[ticker-lookback] wrote {js}")
    print(f"[ticker-lookback] wrote {tl.DAILY / 'ticker_lookback.md'}")
    print(f"[ticker-lookback] phone page dashboard/ticker-lookback/{slug}.html")
    print(f"[ticker-lookback] spreadsheet {xlsx_path}")
    return payload


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tickers", default="",
                    help="comma-separated names, or 'random'")
    ap.add_argument("--random", action="store_true",
                    help="pick 50 stocks with mcap>$100M and avg vol>500K")
    ap.add_argument("--random-n", type=int, default=tl.RANDOM_N)
    ap.add_argument("--seed", default="", help="optional RNG seed for --random")
    ap.add_argument("--from-date", default="", help="optional YYYY-MM-DD")
    ap.add_argument("--to-date", default="", help="optional YYYY-MM-DD")
    args = ap.parse_args()
    seed = args.seed if args.seed else None
    tickers, random_pick = resolve_tickers(
        args.tickers, random_pick=args.random, n=args.random_n,
        asof=args.to_date or None, seed=seed)
    if not tickers:
        raise SystemExit("pass --tickers TEM,ELF or --random")
    run(tickers, from_date=args.from_date or None,
        to_date=args.to_date or None, random_pick=random_pick)


if __name__ == "__main__":
    main()
