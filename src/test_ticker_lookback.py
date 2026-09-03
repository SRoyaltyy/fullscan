"""Ticker-first lookback regression tests against committed artifacts."""
from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src import ticker_lookback as tl
from src import ticker_lookback_run as run


def test_enriched_ab_dates_are_indexed() -> None:
    assert "2026-08-24" in tl.session_dates()
    idx = tl.build_index()
    sess = next(x for x in idx["sessions"] if x["date"] == "2026-08-24")
    assert sess["n_ab"] > 1000


def test_session_dates_skip_weekends() -> None:
    dates = tl.session_dates()
    assert dates
    assert "2026-08-29" not in dates  # Saturday dump
    assert "2026-08-30" not in dates  # Sunday dump
    assert "2026-04-26" not in dates  # Sunday dump
    from datetime import datetime
    for d in dates:
        assert datetime.strptime(d, "%Y-%m-%d").weekday() < 5


def test_any_finviz_name_gets_cards_without_book() -> None:
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-24", to_date="2026-08-25")
    rec = payload["names"][0]
    assert rec["n_sessions"] == 2
    assert rec["n_with_print"] == 2
    assert payload.get("asof") == "09:30_et"
    assert all("prior_finviz" in d["sources"] for d in rec["days"])
    assert all(d.get("asof") == "09:30_et" for d in rec["days"])
    assert all(len(d.get("finviz_factors") or {}) >= 10 for d in rec["days"])
    assert all(len(d.get("ab_factors") or {}) >= 10 for d in rec["days"])


def test_asof_0930_tape_is_prior_session_not_same_day() -> None:
    """Same-day post-close Finviz and same-day book must not color D's boxes."""
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-24", to_date="2026-08-25")
    rec = payload["names"][0]
    idx = tl.build_index()
    by_date = {s["date"]: s for s in idx["sessions"]}
    for day in rec["days"]:
        d = day["date"]
        prior = by_date[d].get("prior_date")
        vintage = day.get("factor_vintage") or {}
        assert vintage.get("asof") == "09:30_et"
        assert vintage.get("prior_date") == prior
        assert vintage.get("vol") == prior
        assert vintage.get("ab") == prior
        assert "finviz" not in day["sources"]
        assert "join" not in day["sources"]
        assert "book" not in day["sources"]
        # Join is D's packet file when weather came from the morning predict.
        if tl.join_packet_ok(d):
            assert vintage.get("join") == d
            assert "packet_join" in day["sources"]
        else:
            assert vintage.get("join") == prior
            if day.get("boxes", {}).get("join") != "missing":
                assert "prior_join" in day["sources"]
        # Morning boxes, when present, are D's pre-open files.
        if "preopen_digest" in day["sources"]:
            assert vintage.get("digest") == d
        if "preopen_judge" in day["sources"]:
            assert vintage.get("judge") == d
        if "preopen_predict" in day["sources"]:
            assert vintage.get("sector") == d
            assert vintage.get("gen") == d

    # 2026-08-24 same-day AAPL RelVol is 0.52 (dead/red). Prior 08-21 is 1.05
    # (yellow). A 09:30 backtest must keep yellow.
    day24 = next(d for d in rec["days"] if d["date"] == "2026-08-24")
    same_day = by_date["2026-08-24"]["finviz"]["AAPL"]
    prior_fv = by_date["2026-08-21"]["finviz"]["AAPL"]
    assert tl._fv_relvol(same_day) < tl.RELVOL_DEAD
    assert tl.RELVOL_DEAD <= tl._fv_relvol(prior_fv) < tl.RELVOL_SPIKE
    assert day24["boxes"]["vol"] == "neutral"
    assert day24["boxes"]["vol"] != "bad"
    # 08-24 has join + weather from the morning predict — use that file,
    # not 08-21's join. Vol/buy still stay on the prior tape.
    assert tl.join_packet_ok("2026-08-24") is True
    assert day24["factor_vintage"]["join"] == "2026-08-24"
    assert "packet_join" in day24["sources"]


def test_join_packet_uses_same_day_ranked_not_prior() -> None:
    """08-12 is the first join file. Clock-strict prior tape has no join;
    the packet recipe should still color join from 08-12's ranked file."""
    assert tl.join_packet_ok("2026-08-12") is True
    assert tl.join_packet_ok("2026-08-11") is False
    payload = run.scan_tickers(["AAPL"], from_date="2026-08-12", to_date="2026-08-12")
    day = payload["names"][0]["days"][0]
    assert day["factor_vintage"]["join"] == "2026-08-12"
    assert "packet_join" in day["sources"]
    assert day["boxes"]["join"] != "missing"
    # No Elite export until 08-13, so 08-12 vol has no prior tape.
    assert day["boxes"]["vol"] == "missing"
    assert day["boxes"]["buy"] != "good"


def test_overnight_buy_is_prior_book_not_same_day_ranker() -> None:
    """ELF printed on the 08-20 13:00 book; at 09:30 that morning it was not a pick."""
    payload = run.scan_tickers(
        ["ELF", "NEE"], from_date="2026-08-20", to_date="2026-08-20")
    by_ticker = {n["ticker"]: n["days"][0] for n in payload["names"]}
    elf = by_ticker["ELF"]
    nee = by_ticker["NEE"]
    assert elf["prior_date"] == "2026-08-19"
    assert elf["boxes"]["buy"] != "good"
    assert "overnight_buy" not in elf["class"]
    assert not elf.get("buy_ranks")
    # NEE was on the 08-19 overnight book, so 08-20 09:30 still sees the pick.
    assert nee["boxes"]["buy"] == "good"
    assert nee["class"] == "overnight_buy"
    assert nee.get("buy_ranks")


def test_events_tilt_uses_dated_file_not_latest() -> None:
    # 2026-08-19 has no dated events file; latest.json is a later scan.
    assert (tl.ROOT / "01_daily" / "events" / "latest.json").exists()
    assert not (tl.ROOT / "01_daily" / "events" / "2026-08-19_events.json").exists()
    assert tl._events_sector_tilt("2026-08-19") == {}
    tilt25 = tl._events_sector_tilt("2026-08-25")
    assert tilt25  # dated 08-25 file has bullish industrials, etc.


def test_phone_html_and_returns() -> None:
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-19", to_date="2026-08-20")
    page = run.render_html(payload)
    assert 'name="viewport"' in page
    assert "🟢 up / positive" in page
    assert "09:30 ET" in page
    assert "<th>Price</th><th>Open</th><th>o→c</th><th>+1d</th><th>+3d</th><th>+1w</th><th>Action</th><th>Cond</th><th>Hall pass</th><th>mid_opp</th><th>Setups</th>" in page
    assert "AAPL" in page
    day0 = payload["names"][0]["days"][0]
    assert day0["forward_returns"]["1d"] is not None
    changes = day0["price_changes"]
    assert changes["price"] is not None
    assert changes["1d"] is not None
    assert changes["1d"] == day0["forward_returns"]["1d"]
    panel = tl._price_panel()
    t = panel["AAPL"]
    i = panel.index.searchsorted(__import__("pandas").Timestamp("2026-08-19"))
    expected = round(100 * (float(t.iloc[i + 1]) / float(t.iloc[i]) - 1), 3)
    assert changes["1d"] == expected
    tones = payload["names"][0]["days"][0]["price_tones"]
    assert tones["1d"] in {"good", "neutral", "bad"}
    assert 'td class="' in page
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "lookback.xlsx"
        run.write_xlsx(payload, p)
        wb = load_workbook(p)
        assert "AAPL" in wb.sheetnames
        heads = [c.value for c in wb["AAPL"][1]]
        assert heads[:7] == ["Date", "Price", "Open", "o→c", "+1d", "+3d", "+1w"]
        assert wb["AAPL"]["D2"].value is not None
        assert wb["AAPL"]["D2"].fill.fgColor.rgb[-6:] in {
            "63BE7B", "FFEB84", "F8696B", "808080"}


def test_random_universe_gates() -> None:
    uni = tl.liquid_universe()
    assert len(uni) >= 10
    path = tl.latest_finviz_path()
    import pandas as pd
    df = pd.read_csv(path, usecols=[
        "Ticker", "Market Cap", "Average Volume",
        "Average True Range", "Price",
    ])
    row = df[df["Ticker"].astype(str).str.upper() == uni[0]].iloc[0]
    assert float(row["Market Cap"]) > 100
    assert float(row["Average Volume"]) > 500
    assert tl.moves_enough(row["Average True Range"], row["Price"])
    assert tl.RANDOM_N == 50
    a = tl.pick_random_tickers(n=tl.RANDOM_N, seed=7)
    b = tl.pick_random_tickers(n=tl.RANDOM_N, seed=7)
    c = tl.pick_random_tickers(n=tl.RANDOM_N, seed=8)
    assert a == b
    assert len(a) == tl.RANDOM_N
    assert len(set(a)) == tl.RANDOM_N
    assert a != c
    ten = tl.pick_random_tickers(n=10, seed=7)
    assert len(ten) == 10
    names, flag = run.resolve_tickers("random", seed=7)
    assert flag is True
    assert names == a


def test_movement_gate_drops_compressed_cvs() -> None:
    """CVS still had range in mid-August; by 09-01 ATR% had compressed."""
    assert tl.atr_pct(2.07, 94.02) == 2.202
    assert tl.moves_enough(2.07, 94.02) is False  # 2.20% < 2.5
    assert tl.moves_enough(2.86, 94.17) is True   # 3.04%
    assert tl.MIN_ATR_PCT == 2.5
    mid = set(tl.liquid_universe(asof="2026-08-17"))
    late = set(tl.liquid_universe(asof="2026-09-01"))
    assert "CVS" in mid
    assert "CVS" not in late
    for mover in ("HIVE", "TSLA", "NVDA"):
        assert mover in late
    quiet = tl.liquid_universe(asof="2026-09-01", min_atr_pct=0)
    assert "CVS" in set(quiet)
    assert len(late) < len(quiet)
    assert len(late) >= 1000


def test_price_tones() -> None:
    assert tl.price_tone(1.2) == "good"
    assert tl.price_tone(-1.2) == "bad"
    assert tl.price_tone(0.1) == "neutral"
    assert tl.price_tone(None) == "missing"


def test_signal_improved_is_strict() -> None:
    worse = {"join": "neutral", "ab": "good"}
    next_worse = {"join": "bad", "ab": "good"}
    assert tl.objectively_better(worse, next_worse) is False

    same = {"join": "neutral", "ab": "good"}
    assert tl.objectively_better(same, same) is False

    better = {"join": "good", "ab": "good"}
    assert tl.objectively_better(worse, better) is True

    # missing on either side is ignored, not treated as a downgrade
    assert tl.objectively_better(
        {"join": "neutral", "ab": "missing"},
        {"join": "good", "ab": "missing"},
    ) is True

    days = [
        {"date": "2026-08-19", "boxes": {"join": "neutral", "ab": "neutral"}},
        {"date": "2026-08-20", "boxes": {"join": "good", "ab": "neutral"}},
        {"date": "2026-08-21", "boxes": {"join": "good", "ab": "bad"}},
    ]
    tl.annotate_signal_improved(days)
    assert days[0]["signal_improved"] is False
    assert days[0]["zero_red"] is True
    assert days[1]["signal_improved"] is True
    assert days[1]["zero_red"] is True
    assert days[2]["signal_improved"] is False
    assert days[2]["signal_alarm"] is True
    assert days[2]["zero_red"] is False
    assert days[2]["condition"]["tone"] in {"good", "neutral", "bad"}

    payload = {
        "generated_at": "t",
        "names": [{"ticker": "TEST", "days": days}],
    }
    page = run.render_html(payload)
    assert 'th class="better clean">🔵⚪ 2026-08-20 09:30 ET</th>' in page
    assert 'th class="clean">⚪ 2026-08-19 09:30 ET</th>' in page
    assert "🚨 2026-08-21 09:30 ET" in page
    assert "+≥3 pts" in page
    assert "purely worse" in page
    md = run.render_md(payload)
    assert "🔵⚪ 2026-08-20" in md
    assert "⚪ 2026-08-19" in md
    assert "🚨 2026-08-21" in md
    assert "| Cond |" in md
    assert "| Hall pass |" in md
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "lookback.xlsx"
        run.write_xlsx(payload, p)
        wb = load_workbook(p)
        assert wb["TEST"]["A3"].fill.fgColor.rgb[-6:] == "5B9BD5"
        assert wb["TEST"]["A2"].fill.fgColor.rgb[-6:] == "FFFFFF"


def test_blue_on_point_jump_and_zero_red() -> None:
    assert tl.box_points({"join": "bad", "ab": "neutral", "gen": "good"}) == 6
    assert tl.zero_red({"join": "neutral", "ab": "good"}) is True
    assert tl.zero_red({"join": "bad", "ab": "good"}) is False
    assert tl.zero_red({"join": "missing"}) is False

    # One cell worse (ab yellow→red) but net points +4 → still blue.
    days = [
        {"date": "2026-08-19", "boxes": {
            "join": "bad", "sector": "bad", "gen": "bad", "ab": "neutral"}},
        {"date": "2026-08-20", "boxes": {
            "join": "good", "sector": "good", "gen": "neutral", "ab": "bad"}},
    ]
    assert tl.objectively_better(days[0]["boxes"], days[1]["boxes"]) is False
    assert tl.point_delta(days[0]["boxes"], days[1]["boxes"]) >= 3
    tl.annotate_signal_improved(days)
    assert days[1]["signal_improved"] is True
    assert days[1]["signal_alarm"] is False
    assert days[1]["zero_red"] is False


def test_alarm_and_condition_majority() -> None:
    assert tl.purely_worse(
        {"join": "good", "ab": "neutral"},
        {"join": "neutral", "ab": "bad"},
    ) is True
    assert tl.purely_worse(
        {"join": "good", "ab": "neutral"},
        {"join": "good", "ab": "good"},
    ) is False
    # Mixed: one better, one worse — not purely worse.
    assert tl.purely_worse(
        {"join": "neutral", "ab": "good"},
        {"join": "good", "ab": "bad"},
    ) is False

    green_major = {k: "good" for k, _ in tl.BOX_COLS}
    red_major = {k: "bad" for k, _ in tl.BOX_COLS}
    mixed = {k: "neutral" for k, _ in tl.BOX_COLS}
    mixed["join"] = mixed["ab"] = mixed["peer"] = "good"
    mixed["news"] = mixed["vol"] = "bad"
    assert tl.general_condition(green_major)["tone"] == "good"
    assert tl.general_condition(red_major)["tone"] == "bad"
    assert tl.general_condition(mixed)["tone"] == "neutral"
    assert tl.general_condition({})["tone"] == "missing"

    days = [
        {"date": "2026-08-19", "boxes": {"join": "good", "ab": "good", "peer": "neutral"}},
        {"date": "2026-08-20", "boxes": {"join": "neutral", "ab": "bad", "peer": "neutral"}},
    ]
    tl.annotate_signal_improved(days)
    assert days[1]["signal_alarm"] is True
    assert days[1]["signal_improved"] is False
    payload = {
        "generated_at": "t",
        "names": [{"ticker": "TEST", "days": days}],
    }
    page = run.render_html(payload)
    assert "🚨 2026-08-20" in page
    assert ">0/2/1<" in page
    assert "Hall pass" in page


def test_color_region_ignores_yellows() -> None:
    green = {k: "good" for k, _ in tl.BOX_COLS}
    red = {k: "bad" for k, _ in tl.BOX_COLS}
    mixed = {k: "neutral" for k, _ in tl.BOX_COLS}
    mixed["join"] = mixed["ab"] = mixed["peer"] = "good"
    mixed["news"] = mixed["vol"] = "bad"
    # Cond is yellow (yellows dominate). G−R = 1 is still mixed.
    assert tl.general_condition(mixed)["tone"] == "neutral"
    assert tl.color_region(mixed)["tone"] == "neutral"
    mixed["heat"] = mixed["buy"] = "good"
    assert tl.color_region(mixed)["tone"] == "good"
    assert tl.color_region(green)["tone"] == "good"
    assert tl.color_region(red)["tone"] == "bad"
    assert tl.color_region({})["tone"] == "missing"
    assert tl.color_region({"join": "good", "ab": "good"})["tone"] == "thin"


def test_tag_context_depends_on_region() -> None:
    # Alarm while the row is still green → first crack.
    days = [
        {"date": "2026-08-19", "boxes": {
            "join": "good", "ab": "good", "peer": "good",
            "vol": "good", "news": "neutral"}},
        {"date": "2026-08-20", "boxes": {
            "join": "good", "ab": "good", "peer": "good",
            "vol": "bad", "news": "neutral"}},
    ]
    tl.annotate_signal_improved(days)
    assert days[1]["signal_alarm"] is True
    assert days[1]["region"]["tone"] == "good"
    assert "first_crack" in days[1]["tag_context"]

    # Blue on a still-red row → turn.
    days = [
        {"date": "2026-08-19", "boxes": {
            "join": "bad", "ab": "bad", "peer": "bad",
            "vol": "bad", "news": "neutral"}},
        {"date": "2026-08-20", "boxes": {
            "join": "neutral", "ab": "bad", "peer": "bad",
            "vol": "bad", "news": "neutral"}},
    ]
    tl.annotate_signal_improved(days)
    assert days[1]["signal_improved"] is True
    assert days[1]["region"]["tone"] == "bad"
    assert "turn" in days[1]["tag_context"]

    # Blue on an already-green row → late.
    days = [
        {"date": "2026-08-19", "boxes": {
            "join": "good", "ab": "neutral", "peer": "good",
            "vol": "good", "news": "neutral"}},
        {"date": "2026-08-20", "boxes": {
            "join": "good", "ab": "good", "peer": "good",
            "vol": "good", "news": "neutral"}},
    ]
    tl.annotate_signal_improved(days)
    assert days[1]["signal_improved"] is True
    assert days[1]["region"]["tone"] == "good"
    assert "late" in days[1]["tag_context"]


def test_last_known_tape_fills_gap_not_same_day() -> None:
    """08-26 has join/predict but no Finviz/AB/peer/book. 08-27 09:30
    still knows 08-25's tape and 08-21's book. Same-day 08-27 Finviz
    and 08-27 book must not color."""
    assert tl.last_predict_date("2026-08-27") == "2026-08-26"
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-27", to_date="2026-08-27")
    day = payload["names"][0]["days"][0]
    vintage = day["factor_vintage"]
    assert vintage["vol"] == "2026-08-25"
    assert vintage["ab"] == "2026-08-25"
    assert vintage["peer"] == "2026-08-21"
    assert vintage["overnight_book"] == "2026-08-21"
    assert day["boxes"]["vol"] != "missing"
    assert day["boxes"]["ab"] != "missing"
    assert day["boxes"]["peer"] != "missing"
    assert day["boxes"]["buy"] != "missing"
    # Evening 08-27 weather is not the morning packet — stay on 08-26 join.
    assert tl.join_packet_ok("2026-08-27") is False
    assert vintage["join"] == "2026-08-26"
    assert "prior_join" in day["sources"]
    # Last morning predict still knowable Thursday 09:30.
    assert day["boxes"]["gen"] != "missing"
    assert vintage["gen"] == "2026-08-26"
    # Same-day 08-27 RelVol is 0.61; last-known 08-25 is 0.08. Both dead,
    # but the vintage must be the earlier file.
    idx = tl.build_index()
    by = {s["date"]: s for s in idx["sessions"]}
    assert tl._fv_relvol(by["2026-08-27"]["finviz"]["AAPL"]) != tl._fv_relvol(
        by["2026-08-25"]["finviz"]["AAPL"])


def test_last_known_does_not_invent_pre_finviz_tape() -> None:
    """No Elite export until 08-13. Do not walk back to April 26."""
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-12", to_date="2026-08-12")
    day = payload["names"][0]["days"][0]
    assert day["boxes"]["vol"] == "missing"
    assert "vol" not in (day.get("factor_vintage") or {})


def test_heat_uses_map_heat_board() -> None:
    """Research captains are empty stubs. Industry board still prints."""
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-26", to_date="2026-08-26")
    day = payload["names"][0]["days"][0]
    assert day["boxes"]["heat"] != "missing"
    assert "preopen_heat" in day["sources"]
    assert day["factor_vintage"].get("heat")


def test_judge_and_digest_use_sector_print() -> None:
    """Random names are not in the 5-ticker judge / 400-name digest sample.
    The day's sector tilt / sector digest still printed."""
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-28", to_date="2026-08-28")
    day = payload["names"][0]["days"][0]
    assert day["boxes"]["judge"] != "missing"
    assert day["boxes"]["digest"] != "missing"
    assert tl.judge_sector_tone(
        {"Technology": "bullish"}, "Technology") == "good"
    assert tl.judge_sector_tone(
        {"Semis/Xlk": "bearish"}, "Technology") == "bad"


def test_lookback_matches_readiness_lattice() -> None:
    """Each day uses the Stock Book readiness row: cameras + coaches + lane."""
    payload = run.scan_tickers(
        ["CRM", "BBAI"], from_date="2026-09-01", to_date="2026-09-01")
    assert payload.get("method") == "stock_book_readiness"
    by = {n["ticker"]: n["days"][0] for n in payload["names"]}
    crm = by["CRM"]
    assert crm["factor_vintage"]["asof"] == "09:30_et"
    assert crm["domains"]["market"] == "bad"
    assert set(crm["domains"]) == {k for k, _ in run.DOMAIN_COLS}
    assert crm["lane"] == "probable"
    assert "HARD_RED" in run._hall_text(crm)
    assert "BUY PROBABLE" in crm["decision"]
    assert crm["labeled"].startswith("join")
    assert crm["labeled_domains"].startswith("mkt")
    assert "yday" in crm["boxes"]
    assert "book" not in crm["sources"]
    bbai = by["BBAI"]
    assert bbai["lane"] == "blocked"
    assert "SELL/AVOID" in bbai["decision"] or "BLOCK BUY" in bbai["decision"]
    md = run.render_md(payload)
    page = run.render_html(payload)
    assert "Stock Book readiness" in md
    assert "| Hall pass |" in md
    assert "| Action |" in md
    assert "| mkt |" in md
    assert "| yΔ |" in md
    assert "BUY PROBABLE" in md
    assert "<th>Hall pass</th>" in page
    assert "<th>mkt</th>" in page
    assert "BUY PROBABLE" in page
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "lookback.xlsx"
        run.write_xlsx(payload, p)
        wb = load_workbook(p)
        heads = [c.value for c in wb["CRM"][1]]
        assert "Hall pass" in heads
        assert "mkt" in heads
        assert "Decision" in heads


def test_lookback_pre_lattice_hall_pass_is_grey() -> None:
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-13", to_date="2026-08-13")
    day = payload["names"][0]["days"][0]
    assert day["lattice_live"] is False
    assert day.get("lane") is None
    from src import gainer_asof
    assert day["lane_label"] == gainer_asof.GREY
    assert day["decision"] == "—"
    assert day["domains"]["market"] in {"good", "bad", "neutral", "missing"}


if __name__ == "__main__":
    test_enriched_ab_dates_are_indexed()
    test_session_dates_skip_weekends()
    test_any_finviz_name_gets_cards_without_book()
    test_asof_0930_tape_is_prior_session_not_same_day()
    test_join_packet_uses_same_day_ranked_not_prior()
    test_overnight_buy_is_prior_book_not_same_day_ranker()
    test_events_tilt_uses_dated_file_not_latest()
    test_phone_html_and_returns()
    test_random_universe_gates()
    test_price_tones()
    test_signal_improved_is_strict()
    test_blue_on_point_jump_and_zero_red()
    test_alarm_and_condition_majority()
    test_color_region_ignores_yellows()
    test_tag_context_depends_on_region()
    test_last_known_tape_fills_gap_not_same_day()
    test_last_known_does_not_invent_pre_finviz_tape()
    test_heat_uses_map_heat_board()
    test_judge_and_digest_use_sector_print()
    test_lookback_matches_readiness_lattice()
    test_lookback_pre_lattice_hall_pass_is_grey()
    print("21 tests passed")
