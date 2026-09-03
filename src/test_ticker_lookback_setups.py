"""Featured lookback setups — matchers, board, and Action surfaces."""
from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook

from src import ticker_lookback as tl
from src import ticker_lookback_run as run
from src import ticker_lookback_setups as setups


def _ids(day):
    return [s["id"] for s in setups.match_day(day)]


def test_first_crack_matcher() -> None:
    day = {
        "signal_alarm": True,
        "signal_improved": False,
        "region": {"tone": "good"},
        "stretch": {"tone": "good"},
        "tag_context": ["first_crack"],
        "boxes": {
            "join": "good", "ab": "good", "peer": "good",
            "vol": "bad", "heat": "good", "judge": "good", "gen": "good",
        },
    }
    assert "tag_context:first_crack" in _ids(day)
    day["signal_alarm"] = False
    day["tag_context"] = []
    assert "tag_context:first_crack" not in _ids(day)


def test_vol_ab_pair_matcher() -> None:
    day = {
        "signal_alarm": False,
        "signal_improved": False,
        "region": {"tone": "good"},
        "stretch": {"tone": "good"},
        "tag_context": [],
        "boxes": {"vol": "good", "ab": "good", "join": "good", "gen": "good"},
    }
    assert "pair:vol=good|ab=good" in _ids(day)
    day["boxes"]["ab"] = "neutral"
    assert "pair:vol=good|ab=good" not in _ids(day)


def test_stretch_blue_neutral_not_region() -> None:
    day = {
        "signal_improved": True,
        "signal_alarm": False,
        "region": {"tone": "good"},
        "stretch": {"tone": "neutral"},
        "tag_context": ["late"],
        "boxes": {"join": "good", "heat": "neutral", "vol": "bad"},
    }
    ids = _ids(day)
    assert "tag_stretch:blue|neutral" in ids
    assert not any(i.startswith("tag_region:") for i in ids)

    day["stretch"] = {"tone": "good"}
    day["region"] = {"tone": "neutral"}
    ids = _ids(day)
    assert "tag_stretch:blue|neutral" not in ids


def test_featured_book_loads_mine_window() -> None:
    book = setups.featured_book()
    window = setups.mine_window()
    assert window["from_date"] == "2026-07-31"
    assert window["to_date"] == "2026-08-27"
    assert window["n_tickers"] >= 2000
    by_id = {s["id"]: s for s in book}
    assert by_id["tag_factor:blue|heat=bad"]["edge_1d"] > 2.5
    assert by_id["tag_context:first_crack"]["verdict"] == "fade"
    assert by_id["tag_stretch:blue|neutral"]["mine_key"] == "blue|neutral"
    assert by_id["tag_stretch:blue|neutral"]["n"] > 200


def test_render_shows_dates_and_effectiveness() -> None:
    payload = run.scan_tickers(
        ["AAPL"], from_date="2026-08-19", to_date="2026-08-27")
    setups.attach_setups(payload)
    payload["generated_at"] = "t"
    md = run.render_md(payload)
    page = run.render_html(payload)
    assert "Setups that paid market-wide" in md
    assert "Setups that paid market-wide" in page
    assert "2026-07-31" in md and "2026-08-27" in md
    assert "1d edge" in md
    assert "| Setups |" in md
    assert "Dates these setups printed (this run)" not in page
    assert payload["setup_hits"], "AAPL in this window should print at least one featured setup"
    hit = payload["setup_hits"][0]
    assert hit["date"]
    assert hit["ticker"] == "AAPL"
    assert str(hit["date"]) in md
    assert 'id="setups"' in page
    assert "<th>Setups</th>" in page
    assert "setup-hit" in page or "setup-chip" in page
    assert 'th class="better' in page or "🔵" in page
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "lookback.xlsx"
        run.write_xlsx(payload, p)
        wb = load_workbook(p)
        assert wb.sheetnames[0] == "Setups"
        assert "AAPL" in wb.sheetnames
        assert wb["Setups"]["A1"].value == "Setups that paid market-wide"
        assert "Setups" in [c.value for c in wb["AAPL"][1]]


def test_overlay_rings_factor_boxes_not_date_cell() -> None:
    days = [
        {"date": "2026-08-19", "boxes": {
            "vol": "good", "ab": "good", "join": "good", "gen": "good",
            "heat": "neutral", "judge": "good"}},
        {"date": "2026-08-20", "boxes": {
            "vol": "good", "ab": "good", "join": "good", "gen": "good",
            "heat": "neutral", "judge": "good"}},
    ]
    payload = {"generated_at": "t", "names": [{"ticker": "TEST", "days": days}]}
    page = run.render_html(payload)
    assert "<th>Setups</th>" in page
    assert 'td class="good setup-hit setup-long"' in page
    assert "vol+AB" in page
    # Date cell stays a single-line mark; chips live in the Setups column.
    assert ">2026-08-20 09:30 ET</th>" in page
    assert "setup-chip" in page
    assert '<div class="setup-hits">' not in page


def test_random_n_is_fifty() -> None:
    assert tl.RANDOM_N == 50
    assert run._slug(["AAA", "BBB", "CCC", "DDD", "EEE"], random_pick=True) == (
        "random-5-aaa-bbb-ccc-ddd"
    )


if __name__ == "__main__":
    test_first_crack_matcher()
    test_vol_ab_pair_matcher()
    test_stretch_blue_neutral_not_region()
    test_featured_book_loads_mine_window()
    test_render_shows_dates_and_effectiveness()
    test_overlay_rings_factor_boxes_not_date_cell()
    test_random_n_is_fifty()
    print("7 setup tests passed")
